import io
import zipfile
from tempfile import NamedTemporaryFile

import dbf
import pandas as pd

from flask import send_file, Response
from flask_login import current_user

from . import db
from .models import (
    Report, Version_report, Sections
)

from sqlalchemy import asc, case, func, desc
from sqlalchemy import func, String
from sqlalchemy.orm import joinedload

def generate_excel_report(version_id):
    current_report = Report.query.filter_by(id=version_id).first()
    
    priority = case(
        {
            "9001": 1,
            "9010": 1,
            "9100": 1
        },
        value=Sections.code_product,
        else_=0
    )

    sections1 = Sections.query.filter_by(id_version=version_id, section_number=1)\
        .order_by(priority, asc(Sections.code_product)).all()

    sections2 = Sections.query.filter_by(id_version=version_id, section_number=2)\
        .order_by(priority, asc(Sections.code_product)).all()

    sections3 = Sections.query.filter_by(id_version=version_id, section_number=3)\
    .order_by(priority, asc(Sections.code_product)).all()
    
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    regular_font_9 = Font(name="Times New Roman", size=9)
    regular_font_9_italic = Font(name="Times New Roman", size=9, italic=True)
    regular_font_10 = Font(name="Times New Roman", size=10)
    regular_font_10_italic = Font(name="Times New Roman", size=10, italic=True)
    
    regular_font_11 = Font(name="Times New Roman", size=11)
    regular_font_11_italic = Font(name="Times New Roman", size=11, italic=True)
    
    regular_font_13 = Font(name="Times New Roman", size=13)
    bold_font_10 = Font(name="Times New Roman", size=10, bold=True)
    bold_font_11 = Font(name="Times New Roman", size=11, bold=True)
    bold_font_13 = Font(name="Times New Roman", size=13, bold=True)
    
    vertical_text = Alignment(horizontal="center", vertical="bottom", textRotation=90, wrap_text=True)
    top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    bottom = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    bottom_left = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_bottom = Side(border_style="thin", color="000000")
    bottom_border = Border(bottom=thin_bottom)


    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def set_cell(ws, row_start, col_start, row_end=None, col_end=None, text="", 
                        font=regular_font_10, row_height=None, alignment=left,
                        merge_direction='both'):  # 'horizontal', 'vertical', 'both', 'none'
        """
        merge_direction: 
        - 'horizontal': объединить только по столбцам (row_end игнорируется)
        - 'vertical': объединить только по строкам (col_end игнорируется)
        - 'both': объединить и по строкам и по столбцам
        - 'none': не объединять
        """
        if merge_direction == 'horizontal':
            row_end = row_start
            if col_end is None:
                col_end = col_start
        elif merge_direction == 'vertical':
            col_end = col_start
            if row_end is None:
                row_end = row_start
        elif merge_direction == 'both':
            if row_end is None:
                row_end = row_start
            if col_end is None:
                col_end = col_start
        else:
            row_end = row_start
            col_end = col_start
        
        if row_start != row_end or col_start != col_end:
            ws.merge_cells(start_row=row_start, start_column=col_start, 
                        end_row=row_end, end_column=col_end)
        
        cell = ws.cell(row=row_start, column=col_start)
        cell.value = text
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        
        if row_height is not None:
            for row in range(row_start, row_end + 1):
                ws.row_dimensions[row].height = row_height
        
        return cell
    
    def page_setttings(ws, print_area):
        # ws.print_area = print_area      
        ws.page_setup.scale = 85
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

    def get_org_name_by_okpo(okpo):
        dop_org_data = [
            ('Брестское областное управление', '100000001000'),
            ('Витебское областное управление', '200000002000'),
            ('Гомельское областное управление', '300000003000'),
            ('Гродненское областное управление', '400000004000'),
            ('Управление г. Минск', '500000005000'),
            ('Минское областное управление', '600000006000'),
            ('Могилевское областное управление', '700000007000'),
            ('Департамент по энергоэффективности', '800000008000'),
        ]
    
        if not okpo or len(str(okpo)) < 4:
            return ""  
        
        okpo_str = str(okpo)
        if len(okpo_str) >= 4:
            fourth_from_end = okpo_str[-4]  
        else:
            return ""
        
        for name, code in dop_org_data:
            if code.endswith(str(fourth_from_end) + "000"):
                return name
        
            return ""

    def title_list(wb, report):
        ws = wb.create_sheet("Титульный лист", 0)
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 15:
                ws.row_dimensions[row].height = 14
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center 
                        
            okpo = report.okpo
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"{org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"областное (городское) управление по надзору за рациональным использованием ТЭР"
                        
            ws.merge_cells("B5:F5")
            ws["B5"].value = org_text
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
            
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(министерство, концерн, государственный комитет)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        # title_first_sign()
        # title_second_sign()
        
        ws.merge_cells("A14:Q15")
        ws["A14"].value = f"ВЕДОМСТВЕННАЯ ОТЧЕТНОСТЬ О НОРМАХ РАСХОДА И (ИЛИ) ПРЕДЕЛЬНЫХ УРОВНЯХ ПОТРЕБЛЕНИЯ ТОПЛИВНО-ЭНЕРГЕТИЧЕСКИХ РЕСУРСОВ ЗА {report.quarter} КВАРТАЛ {report.year} Г."
        ws["A14"].font = bold_font_13
        ws["A14"].alignment = center
                                     
        ws.merge_cells("B16:P17")
        ws["B16"].value = f"{report.organization.full_name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):  # B=2, N=15
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:N18")
        ws["D18"].value = "(наименование юридического лица)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                        
        ws.merge_cells("B20:P20")
        ws["B20"].value = f"ОКПО: {report.okpo}".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Респондент:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:G25")
        ws["E25"].value = "Email"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:G26")
        ws["E26"].value = "ФИО"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "Телефон"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    

        ws.merge_cells("J25:M25")  
        ws["J25"].value = f"{report.user.email}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = left
        
        ws.merge_cells("J26:M26")    
        ws["J26"].value = f"{report.user.fio}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = left    
        
        ws.merge_cells("J27:M27")  
        ws["J27"].value = f"{report.user.telephone}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = left      
                     
        page_setttings(ws, print_area = "A1:O27")
        
        return ws

    def add_sheet(ws, sections, title, unit_header_one_text, unit_header_all_text, nubmerT, nameT):
        ws.title = title
        merged_cells = {
            'A2:I3': f"РАЗДЕЛ {nubmerT}. {nameT}",
            'A4:A5': "Наименование вида продукции (работ услуг)",
            'B4:B5': "Код строки",
            'C4:C5': "Код по ОКЭД",
            'D4:D5': "Единица измерения",
            'E4:E5': "Произведено продукции (работ, услуг) за отчетный период",
            'F4:G4': f"Израсходовано на единицу продукции (работы, услуги) за отчетный период, {unit_header_one_text}",
            'H4:I4': f"Израсходовано на всю произведенную продукцию (работу, услугу) за отчетный период, {unit_header_all_text}",
            'F5': "по утвержденной норме (предельному уровню)",
            'G5': "фактически",
            'H5': "по утвержденной норме (предельному уровню)",
            'I5': "фактически",
            'A6': "A",
            'B6': "Б",
            'C6': "В",
            'D6': "Г",
            'E6': "1",
            'F6': "2",
            'G6': "3",
            'H6': "4",
            'I6': "5"
        }

        font = Font(name='Times New Roman', size=12)
        alignment_header = Alignment(wrap_text=True, vertical='center', horizontal='center')
        alignment_data_left = Alignment(wrap_text=True, vertical='center', horizontal='left')
        alignment_data_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for cell_range, text in merged_cells.items():
            top_left_cell = cell_range.split(':')[0]
            ws[top_left_cell] = text
            ws[top_left_cell].font = font
            ws[top_left_cell].alignment = alignment_header
            ws[top_left_cell].border = border

        for cell_range in merged_cells.keys():
            ws.merge_cells(cell_range)

        column_widths = {
            'A': 50,
            'B': 8,
            'C': 12,
            'D': 12, 
            'E': 12,
            'F': 18,
            'G': 12,
            'H': 18,
            'I': 12,
        }
        row_heights = {
            4: 65,
            5: 65,
            6: 20
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        row_index = 7 
        for section in sections:
            note_text = f' - ({section.note})' if section.note else ''
            product_name = f"{section.product.NameProduct} {note_text}"
            
            ws[f'A{row_index}'] = product_name
            ws[f'B{row_index}'] = section.code_product
            ws[f'C{row_index}'] = section.Oked
            ws[f'D{row_index}'] = section.product.unit.NameUnit
            ws[f'E{row_index}'] = section.produced
            ws[f'F{row_index}'] = section.Consumed_Quota
            ws[f'G{row_index}'] = section.Consumed_Fact
            ws[f'H{row_index}'] = section.Consumed_Total_Quota
            ws[f'I{row_index}'] = section.Consumed_Total_Fact

            
            for col in column_widths.keys():
                cell = ws[f'{col}{row_index}']
                cell.border = border
                cell.font = font
                
                
                if col == 'A':
                    cell.alignment = alignment_data_left
                else:
                    cell.alignment = alignment_data_center

            row_index += 1
            
        page_setttings(ws, print_area="A1:I20")
    
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    wb.active = title_list(wb, current_report)
    
    add_sheet(wb.create_sheet(), sections1, "Раздел 1", 'кг у.т.', 'т у.т.', 1, "ТОПЛИВО")         
    add_sheet(wb.create_sheet(), sections2, "Раздел 2", 'Мкал', 'Гкал', 2, "ТЕПЛОВАЯ ЭНЕРГИЯ")             
    add_sheet(wb.create_sheet(), sections3, "Раздел 3", 'кВтч', 'тыс.кВтч', 3, "ЭЛЕКТРИЧЕСКАЯ ЭНЕРГИЯ")  
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(file_stream, as_attachment=True, download_name=f'{current_report.okpo}_{current_report.year}_{current_report.quarter}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def get_approved_versions(form_data):
    year_filter = form_data.get('year_filter')
    quarter_filter = form_data.get('quarter_filter')

    filters = []
    if year_filter:
        filters.append(Report.year == year_filter)
    if quarter_filter:
        filters.append(Report.quarter == quarter_filter)

    current_user_type = current_user.type
    okpo_digit = str(current_user.organization.okpo)[-4]

    subquery = db.session.query(
        Report.okpo,
        Report.year,
        Report.quarter,
        func.max(Version_report.sent_time).label('max_sent_time')
    ).join(
        Version_report, Version_report.report_id == Report.id
    ).filter(
        Version_report.status == "Одобрен",
        *filters
    ).group_by(
        Report.okpo, Report.year, Report.quarter
    ).subquery()

    query = Version_report.query.options(
        joinedload(Version_report.report),
        joinedload(Version_report.sections).joinedload(Sections.product)
    ).join(Report).join(
        subquery,
        db.and_(
            Report.okpo == subquery.c.okpo,
            Report.year == subquery.c.year,
            Report.quarter == subquery.c.quarter,
            Version_report.sent_time == subquery.c.max_sent_time
        )
    ).filter(
        Version_report.status == "Одобрен",
        *filters
    )

    if not (current_user_type == "Администратор" or (current_user_type == "Аудитор" and okpo_digit == "8")):
        query = query.filter(
            func.substr(func.cast(Report.okpo, String), func.length(Report.okpo) - 3, 1) == okpo_digit
        )

    return query.all()

def create_dbf_zip(versions):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED) as zip_file:
        for version in versions:
            dbf_data = prepare_dbf_data(version)
            dbf_content = create_dbf_file(dbf_data, version.report)
            zip_file.writestr(
                # f'{version.report.okpo}_{version.report.year}_{version.report.quarter}_{version.report.id}.dbf',
                f'{version.report.okpo}_{version.report.year}_{version.report.quarter}.dbf',
                dbf_content
            )
    
    zip_buffer.seek(0)
    return zip_buffer

def prepare_dbf_data(version):
    if not version or not version.report:
        return pd.DataFrame()
    
    report = version.report
    sections = version.sections
    
    data = []
    
    for section in sections:
        product = getattr(section, 'product', None)
        row = create_dbf_row(report, version, section, product)
        data.append(row)
    
    if not data:
        return pd.DataFrame()
    
    df = pd.DataFrame(data)
    df = sort_and_format_dataframe(df)
    return df

def create_dbf_row(report, version, section, product):
    unit_code = str(product.unit.CodeUnit) if (product and hasattr(product, 'unit') and product.unit) else ''
    code_product_str = str(section.code_product) if section.code_product else ''
    
    return {
        'YEAR_': str(report.year) if report.year is not None else '',
        'KVARTAL': str(report.quarter) if report.quarter is not None else '',
        'IDPREDPR': str(report.okpo) if report.okpo is not None else '',
        'DATERECEIV': version.sent_time.strftime('%d.%m.%Y') if version.sent_time else '',
        'EXCEED': '0,000',
        'SECTIONNUM': str(section.section_number) if section.section_number is not None else '',
        'CODEPROD': code_product_str,
        'OKED': str(section.Oked) if section.Oked is not None else '',
        'PRODUCED': format_number(section.produced), 
        'CONSUMEDQ': format_number(section.Consumed_Quota),
        'CONSUMEDF': format_number(section.Consumed_Fact),
        'CONSUMEDQT': format_number(section.Consumed_Total_Quota),
        'CONSUMEDFT': format_number(section.Consumed_Total_Fact),
        'COMMENT1': safe_encode_cp866(str(section.note), max_length=200) if section.note is not None else '',
        'COMMENT2': '',
        'NAMEPROD': safe_encode_cp866(product.NameProduct, max_length=254) if product and hasattr(product, 'NameProduct') and product.NameProduct else '',
        'CODEUNIT': unit_code[:20] if unit_code else '',
        'NAMEORG': safe_encode_cp866(report.organization.full_name, max_length=254) if report.organization.full_name is not None else '',
        '_SECTIONNUM': int(section.section_number) if section.section_number else 0,
        '_CODEPROD': int(section.code_product) if section.code_product and section.code_product.isdigit() else 0,
        '_SORT_PRIORITY': {'9001': 0, '9010': 1, '9100': 2}.get(code_product_str, 3),
    }
    
"""Сортирует и форматирует DataFrame перед экспортом"""
def sort_and_format_dataframe(df):
    df = df.sort_values(by=['_SECTIONNUM', '_SORT_PRIORITY', '_CODEPROD'])
    df = df.drop(columns=['_SECTIONNUM', '_CODEPROD', '_SORT_PRIORITY'])
    df['INDX'] = range(1, len(df) + 1)
    df['INDX'] = df['INDX'].astype(str)
    return df[['INDX'] + [col for col in df.columns if col != 'INDX']]

"""Создает DBF-файл из DataFrame"""
def create_dbf_file(df, report):
    with NamedTemporaryFile(delete=False, suffix='.dbf') as temp_file:
        temp_filename = temp_file.name
        
        table = dbf.Table(
            temp_filename,
            'INDX C(10); YEAR_ C(20); KVARTAL C(20); IDPREDPR C(20); DATERECEIV C(20); '
            'EXCEED C(20); SECTIONNUM C(20); CODEPROD C(20); OKED C(20); PRODUCED C(20); '
            'CONSUMEDQ C(20); CONSUMEDF C(20); CONSUMEDQT C(20); CONSUMEDFT C(20); '
            'COMMENT1 C(200); COMMENT2 C(200); NAMEPROD C(200); CODEUNIT C(20); NAMEORG C(254); ',
            codepage='cp866'
        )
        
        table.open(mode=dbf.READ_WRITE)
        try:
            for _, row in df.iterrows():
                row_dict = row.to_dict()
                for key, value in row_dict.items():
                    if pd.isna(value) or value is None:
                        row_dict[key] = ''
                    elif isinstance(value, str):
                        try:
                            value.encode('cp866', 'strict')
                        except UnicodeEncodeError:
                            row_dict[key] = safe_encode_cp866(value)
                
                table.append(row_dict)
        finally:
            table.close()
        
        with open(temp_filename, 'rb') as f:
            return f.read()
        
"""Форматирует число с запятой в качестве десятичного разделителя"""
def format_number(value, decimal_places=None):
    if value is None or value == '':
        return '0'
    try:
        str_value = str(value).strip()
        str_value = str_value.replace('.', ',').replace(',', '.', 1) if '.' in str_value and ',' in str_value else str_value
        str_value = str_value.replace(',', '.')
        
        num = float(str_value)
        if decimal_places is None:
            if '.' in str_value:
                decimal_places = len(str_value.split('.')[1])
            else:
                decimal_places = 0
        
        if decimal_places == 0:
            return f"{int(round(num))}"
        else:
            formatted = f"{num:.{decimal_places}f}"
            return formatted.replace('.', ',')
            
    except (ValueError, TypeError):
        return '0,000'
    
"""кодирование текста в CP866 с заменой проблемных символов"""
def safe_encode_cp866(text, max_length=None):
    if text is None:
        return ""
    
    text = str(text)
    
    replacements = {
        '\xb3': '3',      # ³ -> 3
        '\xb2': '2',      # ² -> 2
        '\xb9': '1',      # ¹ -> 1
        '\xa3': 'фунт',   # £ -> фунт
        '\xa7': 'S',      # § -> S
        '\xb5': 'мк',     # µ -> мк
        '\xb0': 'град',   # ° -> град
        '\xb1': '+/-',    # ± -> +/-
        '\xf7': '/',      # ÷ -> /
        '\xd7': 'x',      # × -> x
        '\xae': '(R)',    # ® -> (R)
        '\xa9': '(c)',    # © -> (c)
        '\u2122': '(TM)', # ™ -> (TM)
        '\u20ac': 'EUR',  # € -> EUR
        '\u2013': '-',    # – -> -
        '\u2014': '-',    # — -> -
        '\u2018': "'",    # ‘ -> '
        '\u2019': "'",    # ’ -> '
        '\u201c': '"',    # " -> "
        '\u201d': '"',    # " -> "
        '\u2026': '...',  # … -> ...
        '\u2022': '*',    # • -> *
        '\xa0': ' ',      # неразрывный пробел -> пробел
        '\xad': '',       # мягкий перенос
        '\x96': '-',      # – -> -
        '\x97': '-',      # — -> -
        '\x84': '"',      # „ -> "
        '\x93': '"',      # " -> "
        '\x94': '"',      # " -> "
        '\x85': '...',    # … -> ...
    }
    
    for old, new in replacements.items():
        text = text.replace(old, new)
    
    if max_length and len(text) > max_length:
        text = text[:max_length]
    
    try:
        encoded = text.encode('cp866', 'strict')
    except UnicodeEncodeError:

        encoded = text.encode('cp866', 'replace')
    
    return encoded.decode('cp866')

"""Отправляет ZIP-файл"""
def send_zip_file(zip_buffer):
    return Response(
        zip_buffer,
        mimetype='application/zip',
        headers={"Content-Disposition": "attachment;filename=reports_DBF.zip"}
    )