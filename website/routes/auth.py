import os
import re

import io
from io import BytesIO

import zipfile
import random
import string

from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta
import requests
from lxml import etree

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from flask import (
    Blueprint, jsonify, request, flash, redirect, session,
    url_for, send_file, make_response
)

from flask_login import (
    login_user, logout_user, current_user,
    login_required, LoginManager
)

from sqlalchemy import func, desc
from sqlalchemy.orm import joinedload
from sqlalchemy.types import String
from ..export import generate_excel_report
from werkzeug.security import check_password_hash, generate_password_hash
from user_agents import parse

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

from .. import db
from ..models import (
    User, Organization, Report, Version_report, DirUnit,
    DirProduct, Sections, Ticket, Message, UserSession
)

from website.ecp import check_certificate_expiry
from website.session_utils import create_user_session, get_session_token_from_cookie, session_required
from ..time import current_utc_time
from ..email import send_email

auth = Blueprint('auth', __name__)
login_manager = LoginManager()

ZERO_DECIMAL = Decimal('0.00')

@login_manager.user_loader
def load_user(id):
    return User.query.get(int(id))

def last_quarter():
    current_month = current_utc_time().strftime("%m")
    if (current_month == '01' or current_month == '02' or current_month == '03'):
        last_quarter_value = 4
    elif (current_month == '04' or current_month == '05' or current_month == '06'):
        last_quarter_value = 1
    elif (current_month == '07' or current_month == '08' or current_month == '09'):
        last_quarter_value = 2
    else:
        last_quarter_value = 3
    return last_quarter_value

def year_fourMounth_ago():
    months_to_subtract = 4
    new_date = current_utc_time() - timedelta(days=months_to_subtract * 30)
    year_4_months_ago = new_date.year
    return year_4_months_ago

def get_location_info(user_agent_string):
    try:
        ip_response = requests.get("https://api64.ipify.org?format=json")
        ip_response.raise_for_status()
        ip_address = ip_response.json().get("ip")

        location_response = requests.get(f"https://ipinfo.io/{ip_address}/json")
        location_response.raise_for_status()
        location_data = location_response.json()

        user_agent = parse(user_agent_string)
        browser = user_agent.browser.family 
        os = user_agent.os.family
        
        location = location_data.get("city", "Неизвестно") + ", " + location_data.get("region", "Неизвестно") + ", " + location_data.get("country", "Неизвестно")
        ip_address_str = ip_address if ip_address else "Неизвестно"

        return ip_address_str, location, os, browser

    except Exception as e:
        print(f"Ошибка при получении данных о местоположении: {e}")
        return "Неизвестно", "Неизвестно", "Неизвестно", "Неизвестно"  
        
def send_activation_email(email):
    message = gener_password()
    session['activation_code'] = message
    send_email(message, email, 'kod')

def gener_password():
    length=5
    characters = string.digits
    password = ''.join(random.choice(characters) for _ in range(length))
    return password

def parse_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return None

@auth.route('/login', methods=['POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False

        if email and password:
            user = User.query.filter(func.lower(User.email) == func.lower(email)).first()
            if user:
                if check_password_hash(user.password, password):
                    login_user(user, remember=remember)
                    session_token = create_user_session(user.id)
                    resp = make_response(redirect(url_for('views.account')))
                    resp.set_cookie('session_token', session_token, httponly=True)
                    flash('Авторизация прошла успешно.', 'success')
                    return resp
            flash('Неправильный email или пароль.', 'error')
        else:
            flash('Введите данные для авторизации.', 'error')

    return redirect(url_for('views.login'))

@auth.route('/logout')
@login_required
def logout():
    token = request.cookies.get('session_token')
    if token:
        session_obj = UserSession.query.filter_by(session_token=token).first()
        if session_obj:
            db.session.delete(session_obj)
            db.session.commit()

    response = make_response(redirect(url_for('views.login')))
    response.delete_cookie('session_token')
    logout_user()

    flash('Выполнен выход из аккаунта.', 'success')
    return response

@auth.route('/sign', methods=['POST'])
def sign():
    if request.method == 'POST':
        email = request.form.get('email')
        password1 = request.form.get('password1')
        password2 = request.form.get('password2')
        if email and password1:
            if User.query.filter(func.lower(User.email) == func.lower(email)).first():
                flash('Пользователь с таким email уже существует.', 'error')
            elif not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
                flash('Некорректный адрес электронной почты.', 'error') 
            elif password1 != password2:
                flash('Ошибка в подтверждении пароля.', 'error') 
            else:      
                session['temp_user'] = {
                    'email': email,
                    'password': generate_password_hash(password1)
                }
                session.permanent = True
                send_activation_email(email) 
                flash('Проверьте свою почту для активации аккаунта.', 'success')
                return redirect(url_for('views.kod'))
        else:
            flash('Введите данные для регистрации.', 'error')    
    return redirect(url_for('views.sign'))

@auth.route('/kod', methods=['POST'])
def kod():
    if request.method == 'POST':
        input_code = ''.join([
            request.form.get(f'activation_code_{i}', '') for i in range(5)
        ])
        if input_code == session.get('activation_code'):
            new_user = User(
                email=session['temp_user']['email'],
                password=session['temp_user']['password']
            )
            db.session.add(new_user)
            db.session.commit()
            
            remember = True
            
            login_user(new_user, remember=remember)
            session_token = create_user_session(new_user.id)
            resp = make_response(redirect(url_for('views.account')))
            resp.set_cookie('session_token', session_token, httponly=True)
            flash('Аккаунт успешно активирован! Теперь перейдите к заполнению профиля!', 'success')
            return resp
            
            # return redirect(url_for('views.login'))
        else:
            flash('Некорректный код активации.', 'error')
    return redirect(url_for('views.kod'))

@auth.route('/resend-code', methods=['POST'])
def resend_code():
    email = session.get('temp_user', {}).get('email')
    if email:
        new_activation_code = gener_password()
        session['activation_code'] = new_activation_code
        send_email(new_activation_code, email, 'kod')
        return jsonify({'status': 'success'})
    else:
        return jsonify({'status': 'error', 'message': 'Не удалось отправить код повторно.'}), 400

@auth.route('/sessions/clear-others', methods=['POST'])
@login_required
@session_required
def clear_other_sessions():
    current_token = get_session_token_from_cookie()
    other_sessions = UserSession.query.filter(
        UserSession.user_id == current_user.id,
        UserSession.session_token != current_token
    )

    count = other_sessions.count()

    if count == 0:
        flash('Нет других активных сессий.', 'error')
    else:
        other_sessions.delete(synchronize_session=False)
        db.session.commit()
        flash(f'Удалено {count} других сессий.', 'success')
    return redirect(url_for('views.profile_session'))

@auth.route('/add-personal-parametrs', methods=['POST'])
@login_required 
@session_required
def add_personal_parametrs():
    if request.method == 'POST':
        name = request.form.get('name_common', '').strip()
        second_name = request.form.get('second_name_common', '').strip()
        patronymic = request.form.get('patronymic_common', '').strip()
        telephone = request.form.get('telephone_common', '').strip()
        
        id_org = request.form.get('id_org', '').strip()

        if not all([name, second_name, telephone]):
            flash('Заполните все обязательные поля.', 'error')
            return redirect(url_for('views.profile_common'))

        fio = f"{second_name} {name} {patronymic}".strip()
        current_user.fio = fio
        db.session.commit()


        existing_telephone = User.query.filter(User.id != current_user.id, User.telephone == telephone).first()
        if existing_telephone:
            flash('Пользователь с таким номером телефона уже существует.', 'error')
            return redirect(url_for('views.profile_common'))
        current_user.telephone = telephone
        db.session.commit()
        
        if not all([id_org]):
            flash('Выберите предприятие.', 'error')
            return redirect(url_for('views.profile_common'))
        
        organization = Organization.query.filter_by(id=id_org).first()
        if not organization:
            flash('Организация не найдена.', 'error')
            return redirect(url_for('views.profile_common'))
        db.session.commit()

        # existing_userOrg = User.query.filter_by(organization_id=organization.id).first()
        # if existing_userOrg and existing_userOrg.id != current_user.id:
        #     flash('Аккаунт с такой организацией уже существует.', 'error')
        #     return redirect(url_for('views.profile_common'))
        
        current_user.organization_id = organization.id
        
        session.pop('temp_user', None)
        session.pop('activation_code', None)
        
        db.session.commit()
        flash('Данные успешно обновлены.', 'success')
        
    return redirect(url_for('views.profile_common'))

@auth.route('/profile/password', methods=['POST'])
@login_required
def profile_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        conf_new_password = request.form.get('conf_new_password')
        
        if not (old_password and new_password and conf_new_password):
            flash('Введите все поля для смены пароля.', 'error')
            return redirect(url_for('views.profile_password'))

        if not check_password_hash(current_user.password, old_password):
            flash('Неправильный старый пароль.', 'error')
            return redirect(url_for('views.profile_password'))

        if new_password != conf_new_password:
            flash('При подтверждении пароля произошла ошибка.', 'error')
            return redirect(url_for('views.profile_password'))

        user = User.query.filter_by(id=current_user.id).first()
        user.password = generate_password_hash(new_password)
        db.session.commit()

        send_email('Вы успешно изменили свой пароль для входа в учетную запись ErespondentN', current_user.email, 'just_notif')

        UserSession.query.filter_by(user_id=current_user.id).delete()
        db.session.commit()

        logout_user()
        session.clear()

        response = make_response(redirect(url_for('auth.login')))
        response.delete_cookie('session_token')
        flash('Пароль изменён. Выполнен выход из всех устройств.', 'success')
        return response
    return redirect(url_for('views.profile_password'))

@auth.route('/relod-password', methods=['POST'])
def relod_password():
    email = request.form.get('email_relod')
    if not email:
        flash('Пожалуйста, введите свой email, затем нажмите «Забыли пароль?» для восстановления доступа.', 'error')
        return redirect(url_for('views.login'))
    user = User.query.filter(func.lower(User.email) == func.lower(email)).first()
    if user:
        new_password = gener_password()
        hashed_password = generate_password_hash(new_password)

        # user_agent_string = request.headers.get('User-Agent')
        # ip_address, location, os, browser = get_location_info(user_agent_string)
        
        send_email(new_password, email, 'new_pass')
        flash('Новый пароль был отправлен вам на email.', 'success')
        user.password = hashed_password
        db.session.commit()
        return redirect(url_for('views.login'))
    else:
        flash('Пользователя с таким email не существует.', 'error')
        return redirect(url_for('views.login'))

@auth.route('/create-new-report', methods=['POST'])
@login_required 
@session_required
def create_new_report():
    if request.method == 'POST': 
        year =  parse_int(request.form.get('modal_add_year'))
        quarter =  parse_int(request.form.get('modal_add_quarter'))
        
        organization = current_user.organization
        
        has_report = Report.query.filter_by(
            org_id=organization.id, 
            year = year, 
            quarter=quarter,
            user_id = current_user.id).first()
        
        if not has_report:
            new_report = Report(
                okpo=organization.okpo,
                org_id=organization.id,
                year=year,
                quarter=quarter,
                user_id = current_user.id
            )
            db.session.add(new_report)
            db.session.commit()
            new_version_report = Version_report(
                begin_time = current_utc_time(), 
                status = "Заполнение",
                fio = current_user.fio,
                telephone = current_user.telephone,
                email = current_user.email,
                report=new_report
            )
            db.session.add(new_version_report)
            db.session.commit() 

            sections = Sections.query.filter_by(id_version=new_version_report.id).all()
            if not sections:
                id = new_version_report.id
                sections_data = [
                    (id, 332, 9100, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 329, 9010, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 326, 9001, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
      
                    (id, 334, 9100, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 331, 9010, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 328, 9001, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                
                    (id, 333, 9100, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 330, 9010, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 327, 9001, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
        
                ]
                for data in sections_data:
                    section = Sections(
                        id_version=data[0],
                        id_product=data[1],
                        code_product=data[2],
                        section_number=data[3],
                        Oked=data[4],
                        produced=data[5],
                        Consumed_Quota=data[6],
                        Consumed_Fact=data[7],
                        Consumed_Total_Quota=data[8],
                        Consumed_Total_Fact=data[9],
                        total_differents=data[10],
                        note=data[11]
                    )
                    db.session.add(section)
                db.session.commit()
            flash('Добавлен новый отчет.', 'success')
        else:
            flash(f'Отчет {year} года {quarter} квартала уже существует.', 'error')
    return redirect(url_for('views.report_area'))

@auth.route('/change-period-report', methods=['POST'])
@login_required 
@session_required
def change_period_report():
    if request.method == 'POST':
        id = int(request.form.get('modal_change_report_id'))
        year = request.form.get('modal_change_report_year')
        quarter = request.form.get('modal_change_report_quarter')  
         
        current_report = Report.query.filter_by(id=id).first()
        versions = Version_report.query.filter_by(report_id=id).all()

        if current_report:
            existing_report = Report.query.filter_by(
                user_id=current_user.id,
                year=year,
                quarter=quarter
            ).first()

            sent_version_exists = any(version.status == 'Отправлен' for version in versions)
            if sent_version_exists:
                flash('После отправки изменение отчета недоступно.', 'error')
                return redirect(url_for('views.report_area'))
            
            confirmed_version_exists = any(version.status == 'Одобрен' for version in versions)
            if confirmed_version_exists:
                flash('Одобренные отчеты не подлежат редактированию.', 'error')
                return redirect(url_for('views.report_area'))
            
            if existing_report and existing_report.id != id:
                flash('Отчет с таким годом и кварталом уже существует.', 'error')
                return redirect(url_for('views.report_area'))

            current_report.year = year
            current_report.quarter = quarter
            db.session.commit()
            flash('Параметры обновлены.', 'success')
        else:
            flash('Отчет не найден.', 'error')

        return redirect(url_for('views.report_area'))
  
@auth.route('/copy-full-report', methods=['POST'])
@login_required 
@session_required
def copy_full_report():
    if request.method == 'POST':
        try:
            copy_report_id = parse_int(request.form.get('modal_copy_report_id'))
            new_year = parse_int(request.form.get('modal_copy_report_year'))
            new_quarter = parse_int(request.form.get('modal_copy_report_quarter'))
            
            if not all([copy_report_id, new_year, new_quarter]):
                flash('Не все данные заполнены', 'error')
                return redirect(url_for('views.report_area'))
            
            original_report = Report.query.get(copy_report_id)
            if not original_report:
                flash('Исходный отчет не найден', 'error')
                return redirect(url_for('views.report_area'))

            existing_report = Report.query.filter_by(
                org_id=current_user.organization.id,
                year=new_year,
                quarter=new_quarter,
                user_id=current_user.id
            ).first()
            
            if existing_report:
                flash(f'Отчет {new_year} года {new_quarter} квартала уже существует.', 'error')
                return redirect(url_for('views.report_area'))
            
            original_version = Version_report.query.filter_by(
                report_id=copy_report_id
            ).order_by(Version_report.begin_time.desc()).first()
            
            if not original_version:
                flash('Версия исходного отчета не найдена', 'error')
                return redirect(url_for('views.report_area'))
            
            original_sections = Sections.query.filter_by(
                id_version=original_version.id
            ).all()
            
            new_report = Report(
                okpo=current_user.organization.okpo,
                org_id=current_user.organization.id,
                year=new_year,
                quarter=new_quarter,
                user_id=current_user.id
            )
            db.session.add(new_report)
            db.session.flush()
            
            new_version = Version_report(
                begin_time=current_utc_time(),
                status="Заполнение",
                fio=current_user.fio,
                telephone=current_user.telephone,
                email=current_user.email,
                report_id=new_report.id
            )
            db.session.add(new_version)
            db.session.flush()
            
            for section in original_sections:
                new_section = Sections(
                    id_version=new_version.id,
                    id_product=section.id_product,
                    code_product=section.code_product,
                    section_number=section.section_number,
                    Oked=section.Oked,
                    produced=section.produced,
                    Consumed_Quota=section.Consumed_Quota,
                    Consumed_Fact=section.Consumed_Fact,
                    Consumed_Total_Quota=section.Consumed_Total_Quota,
                    Consumed_Total_Fact=section.Consumed_Total_Fact,
                    total_differents=section.total_differents,
                    note=section.note
                )
                db.session.add(new_section)
            
            db.session.commit()
            flash('Отчет успешно скопирован со всеми данными.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при копировании: {str(e)}', 'error')
            
        return redirect(url_for('views.report_area'))

@auth.route('/copy-structure-report', methods=['POST'])
@login_required 
@session_required
def copy_structure_report():
    if request.method == 'POST':
        try:
            copy_report_id = parse_int(request.form.get('modal_copy_report_id'))
            new_year = parse_int(request.form.get('modal_copy_report_year'))
            new_quarter = parse_int(request.form.get('modal_copy_report_quarter'))
            
            if not all([copy_report_id, new_year, new_quarter]):
                flash('Не все данные заполнены', 'error')
                return redirect(url_for('views.report_area'))

            original_report = Report.query.get(copy_report_id)
            if not original_report:
                flash('Исходный отчет не найден', 'error')
                return redirect(url_for('views.report_area'))
            
            existing_report = Report.query.filter_by(
                org_id=current_user.organization.id,
                year=new_year,
                quarter=new_quarter,
                user_id=current_user.id
            ).first()
            
            if existing_report:
                flash(f'Отчет {new_year} года {new_quarter} квартала уже существует.', 'error')
                return redirect(url_for('views.report_area'))
            
            original_version = Version_report.query.filter_by(
                report_id=copy_report_id
            ).order_by(Version_report.begin_time.desc()).first()
            
            if not original_version:
                flash('Версия исходного отчета не найдена', 'error')
                return redirect(url_for('views.report_area'))
            
            original_sections = Sections.query.filter_by(
                id_version=original_version.id
            ).all()
            
            new_report = Report(
                okpo=current_user.organization.okpo,
                org_id=current_user.organization.id,
                year=new_year,
                quarter=new_quarter,
                user_id=current_user.id
            )
            db.session.add(new_report)
            db.session.flush()
            
            new_version = Version_report(
                begin_time=current_utc_time(),
                status="Заполнение",
                fio=current_user.fio,
                telephone=current_user.telephone,
                email=current_user.email,
                report_id=new_report.id
            )
            db.session.add(new_version)
            db.session.flush()
            
            for section in original_sections:
                new_section = Sections(
                    id_version=new_version.id,
                    id_product=section.id_product,
                    code_product=section.code_product,
                    section_number=section.section_number,
                    Oked=section.Oked,
                    produced=ZERO_DECIMAL,
                    Consumed_Quota=ZERO_DECIMAL,
                    Consumed_Fact=ZERO_DECIMAL,
                    Consumed_Total_Quota=ZERO_DECIMAL,
                    Consumed_Total_Fact=ZERO_DECIMAL,
                    total_differents=ZERO_DECIMAL,
                    note=section.note
                )
                db.session.add(new_section)
            
            db.session.commit()
            flash('Отчет успешно скопирован (скопирована только структура).', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при копировании: {str(e)}', 'error')
            
        return redirect(url_for('views.report_area'))

@auth.route('/delete-report/<report_id>', methods=['POST'])
@login_required 
@session_required
def delete_report(report_id):
    if request.method == 'POST':
        current_report = Report.query.filter_by(id = report_id).first()
        versions = Version_report.query.filter_by(report_id = report_id).all()
        tickets = Ticket.query.filter_by(version_report_id = report_id).all()  
        if current_report:   
            sent_version_exists = any(version.status == 'Отправлен' for version in versions)
            if sent_version_exists:
                flash('Отправленный отчет не подлежит удалению.', 'error')
                return redirect(url_for('views.report_area'))  
            confirmed_version_exists = any(version.status == 'Одобрен' for version in versions)
            if confirmed_version_exists:
                flash('Данный отчет не подлежит удалению.', 'error')
                return redirect(url_for('views.report_area'))
            for ticket in tickets:
                db.session.delete(ticket)        
            for version in versions:
                sections = Sections.query.filter_by(id_version = version.id).all()
                for section in sections:
                    db.session.delete(section)
                db.session.delete(version)
            db.session.delete(current_report)
            db.session.commit()
            flash('Отчет удален.', 'success')
        return redirect(url_for('views.report_area'))

@auth.route('/add-section-param', methods=['POST'])
@login_required 
@session_required
def add_section_param():
    if request.method == 'POST':
        current_version_id = request.form.get('current_version')
        add_id_product = request.form.get('add_id_product')
        oked = request.form.get('oked_add')
        produced = request.form.get('produced_add')
        Consumed_Quota = request.form.get('Consumed_Quota_add')
        Consumed_Fact = request.form.get('Consumed_Fact_add')
        Consumed_Total_Quota = request.form.get('Consumed_Total_Quota_add')
        Consumed_Total_Fact = request.form.get('Consumed_Total_Fact_add')
        note = request.form.get('note_add')
        section_number = request.form.get('section_number')
        
        produced = Decimal(produced) if produced else Decimal(0)
        Consumed_Quota = Decimal(Consumed_Quota) if Consumed_Quota else Decimal(0)
        Consumed_Fact = Decimal(Consumed_Fact) if Consumed_Fact else Decimal(0)
        Consumed_Total_Quota = Decimal(Consumed_Total_Quota) if Consumed_Total_Quota else Decimal(0)
        Consumed_Total_Fact = Decimal(Consumed_Total_Fact) if Consumed_Total_Fact else Decimal(0)

        current_product = DirProduct.query.filter_by(id=add_id_product).first()
        if not current_product:
            flash(f'Продукт не найден в справочнике.', 'error')
            return redirect(request.referrer)
        else:
            current_version = Version_report.query.filter_by(id=current_version_id).first()
            product_unit = DirUnit.query.filter_by(IdUnit=current_product.IdUnit).first() 
        
            if current_version:
                if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                    if current_product:
                        proverka_section = Sections.query.filter_by(
                            id_version=current_version_id, 
                            section_number=section_number, 
                            id_product=current_product.id
                        ).first()

                        if proverka_section and not note:
                            flash('Такой вид продукции уже существует — поле "Примечание" обязательно для заполнения.', 'error')
                            return redirect(request.referrer)

                        new_section = Sections(
                            id_version=current_version_id,
                            id_product=current_product.id,
                            code_product=current_product.CodeProduct,
                            section_number=section_number,
                            Oked=oked,
                            produced=produced,
                            Consumed_Quota=Consumed_Quota,
                            Consumed_Fact=Consumed_Fact,
                            Consumed_Total_Quota=Consumed_Total_Quota,
                            Consumed_Total_Fact=Consumed_Total_Fact,
                            total_differents=None,
                            note=note
                        )
                        db.session.add(new_section)
                        db.session.commit()
                        current_section = Sections.query.filter_by(
                            id=new_section.id, 
                            section_number=section_number
                        ).first()

                        if current_section.code_product == "7000":
                            current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                            db.session.commit()
                        else:
                            try:
                                if current_section.produced != 0:
                                    if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                        current_section.Consumed_Fact = round(
                                            (current_section.Consumed_Total_Fact / current_section.produced) * 100, 2)
                                    else:
                                        current_section.Consumed_Fact = round(
                                            (current_section.Consumed_Total_Fact / current_section.produced) * 1000, 2)
                                else:
                                    current_section.Consumed_Fact = 0

                                db.session.commit()
                                if current_section.Consumed_Quota != 0:
                                    if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                        current_section.Consumed_Total_Quota = round(
                                            (current_section.produced * current_section.Consumed_Quota) / 100, 2)
                                    else:
                                        current_section.Consumed_Total_Quota = round(
                                            (current_section.produced * current_section.Consumed_Quota) / 1000, 2)
                                else:
                                    current_section.Consumed_Total_Quota = 0
                                db.session.commit()
                                current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                                db.session.commit()
                            except InvalidOperation as e:
                                flash(f"Ошибка при вычислениях: {e}")

                        specific_codes = ['9001', '9010', '9100']
                        section9001 = Sections.query.filter_by(
                            id_version=current_version_id, 
                            section_number=section_number, 
                            code_product='9001'
                        ).first()
                        aggregated_values = db.session.query(
                            func.sum(Sections.Consumed_Total_Quota),
                            func.sum(Sections.Consumed_Total_Fact),
                            func.sum(Sections.total_differents)
                        ).filter(
                            Sections.id_version == current_version_id,
                            Sections.section_number == section_number,
                            ~Sections.code_product.in_(specific_codes)
                        ).first()

                        if section9001 and aggregated_values:
                            section9001.Consumed_Total_Quota, section9001.Consumed_Total_Fact, section9001.total_differents = aggregated_values
                            db.session.commit()

                        section9010 = Sections.query.filter_by(
                            id_version=current_version_id, section_number=section_number, code_product='9010').first()
                        section9100 = Sections.query.filter_by(
                            id_version=current_version_id, section_number=section_number, code_product='9100').first()

                        if section9100 and section9001 and section9010:
                            section9100.Consumed_Total_Quota = section9001.Consumed_Total_Quota + section9010.Consumed_Total_Quota
                            section9100.Consumed_Total_Fact = section9001.Consumed_Total_Fact + section9010.Consumed_Total_Fact
                            section9100.total_differents = section9001.total_differents + section9010.total_differents
                            db.session.commit()

                        if current_version:
                            current_version.change_time = current_utc_time()
                            current_version.status = "Заполнение"
                            db.session.commit()

                        flash('Продукция была добавлена.', 'success')
                    else:
                        flash('Ошибка при обновлении.', 'error')   
                else:
                    flash('Редактирование отправленного/одобренного отчета недоступно.', 'error')
            else:
                flash('Версия не найдена.', 'error')
    return redirect(request.referrer)
        
@auth.route('/change-section', methods=['POST'])
@login_required 
@session_required
def change_section():
    if request.method == 'POST':
        id_version = request.form.get('current_version')
        id_section = request.form.get('id')
        produced = request.form.get('produced_change')
        Consumed_Quota = request.form.get('Consumed_Quota_change')     
        Consumed_Fact = request.form.get('Consumed_Fact_change')
        Consumed_Total_Quota = request.form.get('Consumed_Total_Quota_change')
        Consumed_Total_Fact = request.form.get('Consumed_Total_Fact_change')
        note = request.form.get('note_change')

        produced = Decimal(produced) if produced else Decimal(0)
        Consumed_Quota = Decimal(Consumed_Quota) if Consumed_Quota else Decimal(0)
        Consumed_Fact = Decimal(Consumed_Fact) if Consumed_Fact else Decimal(0)
        Consumed_Total_Quota = Decimal(Consumed_Total_Quota) if Consumed_Total_Quota else Decimal(0)
        Consumed_Total_Fact = Decimal(Consumed_Total_Fact) if Consumed_Total_Fact else Decimal(0)

        current_version = Version_report.query.filter_by(id=id_version).first() 
        current_section = Sections.query.filter_by(id=id_section).first()
        
        current_product = DirProduct.query.filter_by(id=current_section.id_product).first()
        product_unit = DirUnit.query.filter_by(IdUnit=current_product.IdUnit).first() if current_product else None
        

        if current_version:
            if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                if current_section:
                    if current_section.code_product == "7000":
                        current_section.Consumed_Total_Quota = Consumed_Total_Quota
                        current_section.Consumed_Total_Fact = Consumed_Total_Fact
                        current_section.note = note
                        db.session.commit()
                        
                        current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                        db.session.commit()
                    else:
                        current_section.produced = produced
                        current_section.Consumed_Quota = Consumed_Quota     
                        current_section.Consumed_Total_Fact = Consumed_Total_Fact
                        current_section.note = note 
                        db.session.commit()

                        try:
                            if current_section.produced != 0:
                                if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                    current_section.Consumed_Fact = round(
                                        (current_section.Consumed_Total_Fact / current_section.produced) * 100, 2)
                                else:
                                    current_section.Consumed_Fact = round(
                                        (current_section.Consumed_Total_Fact / current_section.produced) * 1000, 2)
                            else:
                                current_section.Consumed_Fact = 0

                            db.session.commit()
                            if current_section.Consumed_Quota != 0:
                                if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                    current_section.Consumed_Total_Quota = round(
                                        (current_section.produced * current_section.Consumed_Quota) / 100, 2)
                                else:
                                    current_section.Consumed_Total_Quota = round(
                                        (current_section.produced * current_section.Consumed_Quota) / 1000, 2)
                            else:
                                current_section.Consumed_Total_Quota = 0
                            db.session.commit()
                            current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                            db.session.commit()
                        except InvalidOperation as e:
                            flash(f"Ошибка при вычислениях: {e}")



                    if current_version:
                        current_version.change_time = current_utc_time()
                        db.session.commit()

                    specific_codes = ['9001', '9010', '9100']
                    section9001 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product='9001').first()
                    aggregated_values = db.session.query(
                        func.sum(Sections.Consumed_Total_Quota),
                        func.sum(Sections.Consumed_Total_Fact),
                        func.sum(Sections.total_differents)
                    ).filter(
                        Sections.id_version == id_version,
                        Sections.section_number == current_section.section_number,
                        ~Sections.code_product.in_(specific_codes)
                    ).first()

                    if aggregated_values:
                        section9001.Consumed_Total_Quota = aggregated_values[0] or 0
                        section9001.Consumed_Total_Fact = aggregated_values[1] or 0
                        section9001.total_differents = aggregated_values[2] or 0
                        db.session.commit()

                    section9010 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product='9010').first()
                    section9100 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product='9100').first()
                    if section9100 and section9001 and section9010:
                        section9100.Consumed_Total_Quota = (section9001.Consumed_Total_Quota or 0) + (section9010.Consumed_Total_Quota or 0)
                        section9100.Consumed_Total_Fact = (section9001.Consumed_Total_Fact or 0) + (section9010.Consumed_Total_Fact or 0)
                        section9100.total_differents = (section9001.total_differents or 0) + (section9010.total_differents or 0)
                        db.session.commit()
                    current_version.status = "Заполнение"
                    db.session.commit()
                    flash('Параметры обновлены.', 'success')
                else:
                    flash('Ошибка при обновлении.', 'error')   
            else:
                flash('Редактирование отправленного/одобренного отчета недоступно.', 'error')
        else:
            flash('Версия не найдена.', 'error')
        if(current_section.section_number == 1):
            return redirect(url_for('views.report_section', report_type='fuel', id=id_version))
        elif(current_section.section_number == 2):
            return redirect(url_for('views.report_section', report_type='heat', id=id_version))
        elif(current_section.section_number == 3):
            return redirect(url_for('views.report_section', report_type='electro', id=id_version))

@auth.route('/remove_section/<id>', methods=['POST'])
@login_required 
@session_required
def remove_section(id):
    if request.method == 'POST':
        delete_section = Sections.query.filter_by(id=id).first()
        id_version = delete_section.id_version
        section_numberDELsection = delete_section.section_number
        current_version = Version_report.query.filter_by(id=id_version).first() 
        if current_version:
            if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                if delete_section:
                    section9001 = Sections.query.filter_by(id_version=id_version, section_number = section_numberDELsection, code_product = '9001').first()
                    section9001.Consumed_Total_Quota -= delete_section.Consumed_Total_Quota
                    section9001.Consumed_Total_Fact -= delete_section.Consumed_Total_Fact
                    section9001.total_differents -= delete_section.total_differents

                    section9100 = Sections.query.filter_by(id_version=id_version, section_number = section_numberDELsection, code_product = '9100').first()
                    section9100.Consumed_Total_Quota -= delete_section.Consumed_Total_Quota
                    section9100.Consumed_Total_Fact -= delete_section.Consumed_Total_Fact
                    section9100.total_differents -= delete_section.total_differents

                    db.session.delete(delete_section)
                    db.session.commit()
                    flash('Продукция была удалена.', 'success')
                    if current_version:
                        current_version.change_time = current_utc_time()
                        db.session.commit()
                    current_version.status = "Заполнение"
                    db.session.commit()
                else: 
                    flash('Ошибка при удалении', 'error')
            else:
                flash('Редактирование отправленного/одобренного отчета недоступно.', 'error')  
        else:
            flash('Версия не найдена.', 'error')

        match section_numberDELsection:
            case 1:
                return redirect(url_for('views.report_section', report_type='fuel', id=id_version))
            case 2:
                return redirect(url_for('views.report_section', report_type='heat', id=id_version))
            case 3:
                return redirect(url_for('views.report_section', report_type='electro', id=id_version))
            case _:
                return 404

def control_func(id):
    current_version = Version_report.query.filter_by(id=id).first()
    if not current_version:
        flash('Версия отчета не найдена.', 'error')
        return redirect(url_for('views.report_area'))

    id_version = current_version.id


    sections = {
        'fuel': Sections.query.filter_by(id_version=id, section_number=1, code_product='9010').first(),
        'heat': Sections.query.filter_by(id_version=id, section_number=2, code_product='9010').first(),
        'electro': Sections.query.filter_by(id_version=id, section_number=3, code_product='9010').first(),
    }

    if current_version.status == 'Заполнение':
        for key, section in sections.items():
            if section is None or not section.note:
                flash('"Примечание" с кодом строки "9010" обязательно для заполнения.', 'error')
                return redirect(url_for('views.report_section', report_type=key, id=id_version))

        current_version.status = 'Контроль пройден'
        db.session.commit()
        flash('Контроль пройден.', 'successful')
    else:
        flash('Контроль уже был пройден.', 'error')

    return redirect(request.referrer) 

@auth.route('/control-version/<id>', methods=['POST'])
@login_required 
@session_required
def control_version(id):
    if request.method == 'POST':
        return control_func(id)

@auth.route('/agreed-version/<id>', methods=['POST'])
@login_required 
@session_required
def agreed_version(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        if current_version.status == 'Контроль пройден':     
            current_version.status = 'Согласовано'
            db.session.commit()
            flash('Отчет согласован.', 'successful')
        elif current_version.status == 'Согласовано': 
            flash('Отчет уже согласован.', 'succeful')
        else:
            flash('Необходимо пройти контроль.', 'error')
        return redirect(request.referrer)

@auth.route('/send-version/<id>', methods=['POST'])
@login_required 
@session_required
def sent_version(id):
    if request.method == 'POST':
        uploaded_file = request.files.get('certificate')
        current_version = Version_report.query.filter_by(id=id).first()
        
        if current_version.status == 'Отправлен':
            flash('Отчет уже отправлен.', 'error')
            return redirect(request.referrer)

        if current_version.status != 'Согласовано':
            flash('Необходимо согласовать.', 'error')
            return redirect(request.referrer)

        ALLOWED_EXTENSIONS = {'cer'}
        def allowed_file(filename):
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

        if not uploaded_file:
            flash('Файл сертификата обязателен.', 'error')
            return redirect(request.referrer)

        if not allowed_file(uploaded_file.filename):
            flash('Неверный формат файла. Загрузите файл в формате .cer.', 'error')
            return redirect(request.referrer)

        if not check_certificate_expiry(uploaded_file):
            flash('Срок действия сертификата истёк или файл некорректен.', 'error')
            return redirect(request.referrer)

        current_version.status = 'Отправлен'
        if current_version.sent_time is None:
            current_version.sent_time = current_utc_time()
        db.session.commit()

        flash('Отчет отправлен.', 'successful')
        return redirect(request.referrer)

@auth.route('/cancel-sent-version/<id>', methods=['POST'])
@login_required 
@session_required
def cancle_sent_version(id):
    from ..report import cancel_sending
    return cancel_sending(id)

@auth.route('/change-category-report', methods=['POST'])
@login_required 
@session_required
def change_category_report():
    action = request.form.get('action')
    report_id = request.form.get('reportId')
    status_itog = None
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=report_id).first()
        recipient_user = User.query.filter_by(email=current_version.email).first()

        if current_version is not None: 
            user = User.query.filter_by(email=current_version.email).first()
            if not current_version.hasNot and action != 'to_download':
                flash('Необходимо уточнить о каких ошибках идет речь.', 'error')
                return redirect(url_for('views.audit_report', id=current_version.id, tickets_cont='true'))
            if action == 'not_viewed':
                status_itog = 'Отправлен'
            elif action == 'remarks':
                status_itog = 'Есть замечания'
            elif action == 'to_download':
                status_itog = 'Одобрен'
                ticket_message = Ticket(
                    note="Ошибок нет, отчет передан в следующую стадию проверки.",
                    luck=True,
                    version_report_id=current_version.id
                )
                db.session.add(ticket_message)
            elif action == 'to_delete':
                status_itog = 'Готов к удалению'
            else:
                flash('Неизвестное действие.', 'error')
                return redirect(request.referrer) 
            
            current_version.hasNot = False
            current_version.status = status_itog
            current_version.audit_time = current_utc_time()
            db.session.commit()

            user_message = Message(
                text = f"Статус вашего отчета был изменен на {status_itog}. Дополнительные сведения можно просмотреть в квитанции.",
                sender_id = current_user.id,          
                recipient_id = recipient_user.id
            )
            db.session.add(user_message)
            db.session.commit()
    
            send_email(status_itog, user.email, 'status')

            flash(f'Статус отчета был изменен.', 'success')
            return redirect(request.referrer) 
        else:
            flash('Отчет не найден.', 'error')
            return "Version not found", 404
        
@auth.route('/rollbackreport/<id>', methods=['POST'])
@login_required 
@session_required
def rollbackreport(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        recipient_user = User.query.filter_by(email=current_version.email).first()    
        if current_version:
            if current_version.status != 'Отправлен':
                if isinstance(current_version.audit_time, datetime):
                    audit_time = current_version.audit_time
                else:
                    audit_time = datetime.combine(current_version.audit_time, datetime.min.time())

                if audit_time + timedelta(days=30) <= current_utc_time():
                    flash('Прошло больше 30-ти дней, статус отчета изменить нельзя.', 'error')
                else: 
                    current_version.status = "Отправлен"
                    current_version.hasNot = False
                 
                    user_message = Message(
                        text = f"Статус отчета был изменен.",
                        sender_id = recipient_user.id,    
                        recipient_id = recipient_user.id      
                    )
                    db.session.add(user_message)
                    db.session.commit()
                    
                    ticket_message = Ticket(
                        note="Возвращён в исходное состояние.",
                        luck=False,
                        version_report_id=current_version.id
                    )
                    db.session.add(ticket_message)
                    db.session.commit()
                    flash('Статус отчёта был изменён на «Непросмотренный».', 'success')
            else:
                flash('Статус отчёта уже установлен как «Непросмотренный».', 'error')
            return redirect(request.referrer)
        else:
            flash('Отчет не найден.', 'error')
    return redirect(request.referrer)

@auth.route('/send-comment', methods=['POST'])
@login_required 
@session_required
def send_comment():
    if request.method == 'POST':
        version_id = request.form.get('version_id')
        text = request.form.get('text')

        if not text or text.strip() == '':
            flash("Необходимо ввести текст.", "error")
            return redirect(request.referrer)
        
        cleaned_text = ' '.join(text.split())
        current_version = Version_report.query.filter_by(id=version_id).first()
        current_report = Report.query.get_or_404(version_id)
        
        if current_version:
            new_comment = Ticket(
                note = cleaned_text,
                version_report_id = current_version.id
            )
            db.session.add(new_comment)
            current_version.hasNot = True
            db.session.commit()
            flash('Сообщение отправлено, теперь можно сменить статус.', 'success')
        
            status_mapping = {
                'Отправлен': 'not_viewed',
                'Есть замечания': 'remarks',
                'Одобрен': 'to_download',
                'Готов к удалению': 'to_delete'
            }

            current_status = current_version.status
            url_status = status_mapping.get(current_status, 'all_reports')

            return redirect(url_for('views.audit_area', 
                                status=url_status,
                                year=current_report.year,
                                quarter=current_report.quarter))
        else:
            flash('Отчет не найден.', 'error') 
            return redirect(request.referrer)

            
@auth.route('/export-table', methods=['POST'])
@login_required 
@session_required
def export_table():
    version_id = int(request.form.get('version_id'))
    return generate_excel_report(version_id)

@auth.route('/export-version/<id>', methods=['POST'])
@login_required 
@session_required
def export_version(id):
    return generate_excel_report(id)

@auth.route('/print-version-tickets', methods=['POST'])
@login_required 
@session_required
def print_version_tickets():
    if request.method == 'POST':
        version_id = request.form.get('version_id')
        
        tickets = Ticket.query.filter_by(version_report_id=version_id).all()
        
        if not tickets:
            flash("Квитанции не найдены.", "error")
            return redirect(request.referrer)

        version_report = tickets[0].version_report
        report = version_report.report

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)

        font_path_bold = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'fonts', 'Montserrat-Bold.ttf')
        font_path_regular = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'fonts', 'Montserrat-Regular.ttf')
        pdfmetrics.registerFont(TTFont('MontserratBold', font_path_bold))
        pdfmetrics.registerFont(TTFont('MontserratRegular', font_path_regular))

        c.setFont("MontserratBold", 16)
        c.drawString(100, 780, "Квитанции по отчету")
        
        c.setFont("MontserratRegular", 12)
        c.drawString(100, 760, f"ОКПО: {report.okpo}")
        c.drawString(100, 740, f"Год: {report.year}, Квартал: {report.quarter}")
        c.drawString(100, 720, f"Всего квитанций: {len(tickets)}")
        
        y_position = 680

        for ticket in tickets:
            if y_position < 100:
                c.showPage()
                c.setFont("MontserratBold", 16)
                c.drawString(100, 780, "Квитанции по отчету (продолжение)")
                c.setFont("MontserratRegular", 12)
                c.drawString(100, 760, f"ОКПО: {report.okpo}")
                c.drawString(100, 740, f"Год: {report.year}, Квартал: {report.quarter}")
                y_position = 700

            c.setFont("MontserratBold", 12)
            c.drawString(100, y_position, f"Дата обработки: {ticket.begin_time.strftime('%Y-%m-%d %H:%M')}")
            y_position -= 20
            
            c.setFont("MontserratBold", 12)
            c.drawString(100, y_position, "Результат:")
            c.setFont("MontserratRegular", 12)
            result = "Отчет принят в обработку" if ticket.luck else "Отчет не принят в обработку"
            c.drawString(180, y_position, result)
            y_position -= 20
            
            if not ticket.luck:
                c.setFont("MontserratBold", 12)
                c.drawString(100, y_position, "Причина:")
                c.setFont("MontserratRegular", 12)
                c.drawString(180, y_position, ticket.note if ticket.note else "")
                y_position -= 20
            
            y_position -= 20

        c.save()
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=' + f"kvitancii_{report.okpo}_{report.year}_{report.quarter}.pdf"
        
        return response

@auth.route('/exportDBF', methods=['POST'])
@login_required 
@session_required
def exportDBF():
    if request.method == 'POST':
        from ..export import create_dbf_zip, get_approved_versions, send_zip_file
        try:
            versions = get_approved_versions(request.form)
            if not versions:
                flash('Нет одобренных отчетов для выбранных фильтров.', 'error')
                return redirect(request.referrer)

            zip_buffer = create_dbf_zip(versions)
            return send_zip_file(zip_buffer)
        except Exception as e:
            flash(f'Ошибка при экспорте DBF: {str(e)}', 'error')
            return redirect(request.referrer)

from collections import defaultdict
from lxml import etree
from collections import defaultdict

def create_xml_for_version(version):
    report = version.report
    organization = report.organization

    root = etree.Element("REPOPT_ROOT")
    report_el = etree.SubElement(root, "REPORT")

    etree.SubElement(report_el, "DESCRIPTION_REPORT", attrib={
        "CODE": str(report.okpo or ""),
        "NO": organization.full_name or "",
        "CODE_ENT": organization.full_name or "",
        "OKPO": str(report.okpo or ""),
        "UNP": str(organization.ynp or ""),
        "YEAR": str(report.year),
        "QUARTER": str(report.quarter),
        "PHASE": "1",
        "ISYEAR": "false",
        "TYPE": "2",
        "ISREADONLY": "false",
        "FIO": version.fio or "",
        "PHONE": version.telephone or "",
        "EMAIL": version.email or "",
        "SV": "1.2.7.2"
    })

    row_data_el = etree.SubElement(report_el, "ROW_DATA")
    sections_el = etree.SubElement(row_data_el, "SECTIONS")

    sections_by_number = defaultdict(list)
    for section in version.sections:
        sections_by_number[section.section_number].append(section)

    special_codes_order = ["9001", "9010", "9100"]

    for section_number, group in sections_by_number.items():
        for code in special_codes_order:
            for section in group:
                product = section.product
                if (product.CodeProduct or "") == code:
                    unit = product.unit
                    etree.SubElement(sections_el, "row", attrib={
                        "SectionNumber": str(section.section_number),
                        "SortIndex": code,
                        "CodeProduct": code,
                        "NameProduct": product.NameProduct or "",
                        "OKED": section.Oked or "",
                        "CodeUnit": unit.CodeUnit if unit else "",
                        "Produced": str(section.produced or 0),
                        "ConsumedQ": str(section.Consumed_Quota or 0),
                        "ConsumedF": str(section.Consumed_Fact or 0),
                        "ConsumedQT": str(section.Consumed_Total_Quota or 0),
                        "ConsumedFT": str(section.Consumed_Total_Fact or 0),
                        "Comment": section.note or ""
                    })

        sort_index_counter = 1
        for section in group:
            product = section.product
            code_product = product.CodeProduct or ""
            if code_product not in special_codes_order:
                unit = product.unit
                etree.SubElement(sections_el, "row", attrib={
                    "SectionNumber": str(section.section_number),
                    "SortIndex": str(sort_index_counter),
                    "CodeProduct": code_product,
                    "NameProduct": product.NameProduct or "",
                    "OKED": section.Oked or "",
                    "CodeUnit": unit.CodeUnit if unit else "",
                    "Produced": str(section.produced or 0),
                    "ConsumedQ": str(section.Consumed_Quota or 0),
                    "ConsumedF": str(section.Consumed_Fact or 0),
                    "ConsumedQT": str(section.Consumed_Total_Quota or 0),
                    "ConsumedFT": str(section.Consumed_Total_Fact or 0),
                    "Comment": section.note or ""
                })
                sort_index_counter += 1

    return etree.tostring(root, pretty_print=True, encoding='utf-8')

@auth.route('/exportXML', methods=['POST'])
@login_required 
@session_required
def exportXML():
    year_filter = request.form.get('year_filter')
    quarter_filter = request.form.get('quarter_filter')

    filters = []
    if year_filter:
        filters.append(Report.year == year_filter)
    if quarter_filter:
        filters.append(Report.quarter == quarter_filter)

    current_user_type = current_user.type
    okpo_digit = str(current_user.organization.okpo)[-4]

    if current_user_type == "Администратор" or (current_user_type == "Аудитор" and okpo_digit == "8"):
        versions = Version_report.query.options(
            joinedload(Version_report.report),
            joinedload(Version_report.sections).joinedload(Sections.product)
        ).join(Report).filter(
            Version_report.status == "Одобрен",
            *filters
        ).all()
    else:
        versions = Version_report.query.options(
            joinedload(Version_report.report),
            joinedload(Version_report.sections).joinedload(Sections.product)
        ).join(Report).filter(
            Version_report.status == "Одобрен",
            func.substr(func.cast(Report.okpo, String), func.length(Report.okpo) - 3, 1) == okpo_digit,
            *filters
        ).all()
    
    if not versions:
        flash('Нет одобренных отчетов для выбранных фильтров.', 'error')
        return redirect(request.referrer) 

    mem_zip = io.BytesIO()

    with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for version in versions:
            report = version.report
            xml_data = create_xml_for_version(version)
            file_name = f"{report.okpo}_{report.year}_{report.quarter}.xml"
            zipf.writestr(file_name, xml_data)
    mem_zip.seek(0)

    return send_file(
        mem_zip,
        as_attachment=True,
        download_name="reports_XML.zip",
        mimetype="application/zip"
    )

@auth.route('/sent_for_admin', methods=['POST'])
@login_required 
@session_required
def sent_for_admin():
    if request.method == 'POST':
        text = request.form.get('text')
        if text:
            admins = User.query.filter_by(type = "Администратор").all()
            if admins:
                for i in admins:
                    new_message = Message(
                        sender_id = current_user.id,
                        text = text,
                        recipient_id = i.id
                    )
                    db.session.add(new_message)
                db.session.commit()
                    
                flash('Сообщение отправлено.', 'succes')
                admin_user = os.getenv('adminemail3')
                if admin_user:
                    try:
                        send_email(text, admin_user, 'just_notif')
                    except Exception as e:
                        auth.logger.error(f"Ошибка отправки email: {str(e)}")
            else:
                flash('Администраторов нет.', 'error')
        else:
            flash('Нельзя отравить пустое сообщение.', 'error')
    return redirect(url_for('views.beginPage'))






def get_organizations_with_reports_excel_xlsx(year: int, quarter: int, statuses: list) -> bytes:
    status_filter = 'all_reports' if not statuses else statuses[0] if len(statuses) == 1 else 'all_reports'
    
    from .views import get_reports_by_status
    reports = get_reports_by_status(status_filter, year, quarter)
    
    if not reports:
        return None

    organizations_data = set()
    for report in reports:
        valid_versions = [v for v in report.versions if not statuses or v.status in statuses]
        if valid_versions:
            latest_version = max(valid_versions, key=lambda x: x.sent_time or datetime.min)
            organizations_data.add((
                report.organization.okpo,
                report.organization.full_name,
                latest_version.sent_time
            ))

    if not organizations_data:
        return None

    records = sorted(organizations_data, key=lambda x: x[0] or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Организации"

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="C6EFCE")
    title_font = Font(bold=True, size=12)
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_row = 3
    start_col = 2

    title_text = f"Список предприятий, представивших отчеты по форме Сведения о нормах в электронном виде на {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws.merge_cells(start_row=start_row - 1, start_column=start_col, end_row=start_row - 1, end_column=start_col + 3)
    title_cell = ws.cell(row=start_row - 1, column=start_col, value=title_text)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row - 1].height = 45

    for col in range(start_col, start_col + 4):
        cell = ws.cell(row=start_row - 1, column=col)
        cell.border = thin_border

    headers = ['Код предприятия (ОКПО)', 'Наименование предприятия', 'Дата поступления', 'Примечание']
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    for row_offset, (okpo, full_name, sent_time) in enumerate(records, start=1):
        row_data = [okpo or '', full_name or '', sent_time or '', '']
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(
                row=start_row + row_offset,
                column=start_col + col_offset,
                value=value
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            if col_offset == 2 and sent_time:
                cell.number_format = 'YYYY-MM-DD'

    for i in range(len(headers)):
        col_letter = get_column_letter(start_col + i)
        max_length = len(headers[i])
        for j in range(1, len(records) + 1):
            val = ws.cell(row=start_row + j, column=start_col + i).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@auth.route('/load_org_stat', methods=['POST'])
@login_required 
@session_required
def load_org_stat():
    year_filter = request.form.get('modal_add_year')
    quarter_filter = request.form.get('modal_add_quarter')

    if not year_filter or not quarter_filter:
        flash('Не указан год или квартал.', 'error')
        return redirect(request.referrer)

    if current_user.type not in ["Администратор", "Аудитор"]:
        flash('У вас нет доступа к отчетам.', 'error')
        return redirect(request.referrer)

    allowed_statuses = ["Отправлен", "Одобрен", "Есть замечания", "Готов к удалению"]
    file_data = get_organizations_with_reports_excel_xlsx(
        int(year_filter), int(quarter_filter), allowed_statuses
    )

    if not file_data:
        flash('Нет организаций с отправленными отчётами за выбранный период.', 'error')
        return redirect(request.referrer)

    response = make_response(file_data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=orgs_{year_filter}_{quarter_filter}.xlsx'
    return response

def create_test_users(userpassword):
    hashed_pass = generate_password_hash(userpassword)
    users_data = [
        ('testuser1BrestReg@gmail.com', 'Григорьев Антон Олегович', 'Пусто', hashed_pass, 3),
        ('testuser2BrestReg@gmail.com', 'Иванов Алексей Дмитриевич', 'Пусто', hashed_pass, 4),
        ('testuser1VitebskReg@gmail.com', 'Петров Николай Сергеевич', 'Пусто', hashed_pass, 1170),
        ('testuser2VitebskReg@gmail.com', 'Кузнецов Андрей Валерьевич', 'Пусто', hashed_pass, 1171),
        ('testuser1GomelReg@gmail.com', 'Фёдоров Дмитрий Иванович', 'Пусто', hashed_pass, 2229),
        ('testuser2GomelReg@gmail.com', 'Морозов Сергей Александрович', 'Пусто', hashed_pass, 2230),
        ('testuser1GrodnoReg@gmail.com', 'Васильев Игорь Михайлович', 'Пусто', hashed_pass, 3431),
        ('testuser2GrodnoReg@gmail.com', 'Новиков Евгений Павлович', 'Пусто', hashed_pass, 3432),
        ('testuser1MinskReg@gmail.com', 'Мельников Артём Владимирович', 'Пусто', hashed_pass, 4674),
        ('testuser2MinskReg@gmail.com', 'Орлов Роман Викторович', 'Пусто', hashed_pass, 4675),
        ('testuser1MogilevReg@gmail.com', 'Смирнов Константин Петрович', 'Пусто', hashed_pass, 6607),
        ('testuser2MogilevReg@gmail.com', 'Козлов Денис Сергеевич', 'Пусто', hashed_pass, 6608),
        ('testuser1Minsk@gmail.com', 'Зайцев Павел Андреевич', 'Пусто', hashed_pass, 7493),
        ('testuser2Minsk@gmail.com', 'Соловьёв Владимир Николаевич', 'Пусто', hashed_pass, 6899),
    ]

    current_time = datetime.utcnow()
    current_year = current_time.year
    current_quarter = 1

    created_version_ids = []

    for user_data in users_data:
        user = User.query.filter_by(email=user_data[0]).first()
        if not user:
            user = User(
                email=user_data[0],
                fio=user_data[1],
                telephone=user_data[2],
                password=user_data[3],
                organization_id=user_data[4],
            )
            db.session.add(user)
            db.session.flush()

        organization = Organization.query.get(user.organization_id)
        if not organization:
            continue

        # 2 отчёта: текущий год и следующий квартал
        for quarter in (current_quarter, current_quarter + 1):
            new_report = Report(
                okpo=organization.okpo,
                org_id=organization.id,
                year=current_year,
                quarter=quarter,
                user_id=user.id
            )
            db.session.add(new_report)
            db.session.flush()

            new_version_report = Version_report(
                begin_time=current_time,
                status="Заполнение",
                fio=user.fio,
                telephone=user.telephone,
                email=user.email,
                report=new_report
            )
            db.session.add(new_version_report)
            db.session.flush()

            created_version_ids.append(new_version_report.id)

            sections = Sections.query.filter_by(id_version=new_version_report.id).all()
            if not sections:
                id_version = new_version_report.id
                sections_data = [
                    (id_version, 326, 9100, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('33.00'), Decimal('411.00'), Decimal('378.00'), ''),
                    (id_version, 329, 9010, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), 'какая-то информация'),
                    (id_version, 332, 9001, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('33.00'), Decimal('411.00'), Decimal('378.00'), ''),

                    (id_version, 6, '0026', 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('33.00'), Decimal('411.00'), Decimal('378.00'), ''),

                    (id_version, 327, 9100, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id_version, 330, 9010, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), 'какая-то информация'),
                    (id_version, 333, 9001, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id_version, 328, 9100, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id_version, 331, 9010, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), 'какая-то информация'),
                    (id_version, 334, 9001, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                ]
                for data in sections_data:
                    section = Sections(
                        id_version=data[0],
                        id_product=data[1],
                        code_product=data[2],
                        section_number=data[3],
                        Oked=data[4],
                        produced=data[5],
                        Consumed_Quota=data[6],
                        Consumed_Fact=data[7],
                        Consumed_Total_Quota=data[8],
                        Consumed_Total_Fact=data[9],
                        total_differents=data[10],
                        note=data[11]
                    )
                    db.session.add(section)

    db.session.commit()

    versions_to_update = Version_report.query.filter(Version_report.id.in_(created_version_ids)).all()
    for version in versions_to_update:
        version.status = 'Отправлен'
        version.sent_time = current_utc_time()
    db.session.commit()

    return "Пользователи и два отчёта на каждый год созданы"

def delete_test_users():
    emails = [
        'testuser1BrestReg@gmail.com',
        'testuser2BrestReg@gmail.com',
        'testuser1VitebskReg@gmail.com',
        'testuser2VitebskReg@gmail.com',
        'testuser1GomelReg@gmail.com',
        'testuser2GomelReg@gmail.com',
        'testuser1GrodnoReg@gmail.com',
        'testuser2GrodnoReg@gmail.com',
        'testuser1MinskReg@gmail.com',
        'testuser2MinskReg@gmail.com',
        'testuser1MogilevReg@gmail.com',
        'testuser2MogilevReg@gmail.com',
        'testuser1Minsk@gmail.com',
        'testuser2Minsk@gmail.com',
    ]
    
    users = User.query.filter(User.email.in_(emails)).all()
    for user in users:
        db.session.delete(user)
    db.session.commit()
    return "Пользователи удалены"

@auth.route('/create-test-users', methods=['POST'])
@login_required 
@session_required
def create_test_users_rout():
    return create_test_users(userpassword = '1111')

@auth.route('/delete-test-users', methods=['POST'])
@login_required 
@session_required
def delete_test_users_rout():
    return delete_test_users()