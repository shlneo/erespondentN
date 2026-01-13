import os
import re
import io
import zipfile
import random
import string
import smtplib
from io import BytesIO
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from tempfile import NamedTemporaryFile
import dbf
import pandas as pd
import requests
import xml.etree.ElementTree as ET
from lxml import etree

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from flask import (
    Blueprint, jsonify, request, flash, redirect, session,
    url_for, send_file, Response, make_response
)

from flask_login import (
    login_user, logout_user, current_user,
    login_required, LoginManager
)

from sqlalchemy import func, desc
from sqlalchemy.orm import joinedload
from sqlalchemy.types import String
from werkzeug.security import check_password_hash, generate_password_hash
from user_agents import parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from dotenv import load_dotenv
from . import db
from .models import (
    User, Organization, Report, Version_report, DirUnit,
    DirProduct, Sections, Ticket, Message, UserSession
)

from website.ecp import check_certificate_expiry
from website.session_utils import create_user_session, get_session_token_from_cookie, session_required

from .time_for_app import current_utc_time

auth = Blueprint('auth', __name__)
login_manager = LoginManager()

@login_manager.user_loader
def load_user(id):
    return User.query.get(int(id))

def last_quarter():
    current_month = current_utc_time().strftime("%m")
    if (current_month == '01' or current_month == '02' or current_month == '03'):
        last_quarter_value = 4
    elif (current_month == '04' or current_month == '05' or current_month == '06'):
        last_quarter_value = 1
    elif (current_month == '07' or current_month == '08' or current_month == '09'):
        last_quarter_value = 2
    else:
        last_quarter_value = 3
    return last_quarter_value

def year_fourMounth_ago():
    months_to_subtract = 4
    new_date = current_utc_time() - timedelta(days=months_to_subtract * 30)
    year_4_months_ago = new_date.year
    return year_4_months_ago

def get_location_info(user_agent_string):
    try:
        ip_response = requests.get("https://api64.ipify.org?format=json")
        ip_response.raise_for_status()
        ip_address = ip_response.json().get("ip")

        location_response = requests.get(f"https://ipinfo.io/{ip_address}/json")
        location_response.raise_for_status()
        location_data = location_response.json()

        user_agent = parse(user_agent_string)
        browser = user_agent.browser.family 
        os = user_agent.os.family
        
        location = location_data.get("city", "Неизвестно") + ", " + location_data.get("region", "Неизвестно") + ", " + location_data.get("country", "Неизвестно")
        ip_address_str = ip_address if ip_address else "Неизвестно"

        return ip_address_str, location, os, browser

    except Exception as e:
        print(f"Ошибка при получении данных о местоположении: {e}")
        return "Неизвестно", "Неизвестно", "Неизвестно", "Неизвестно"

def send_email(message_body, recipient_email, email_type, location=None, device=None, browser=None, ip_address=None):
    sender = os.getenv('EMAILNAME')
    password = os.getenv('EMAILPASS')
    
    base_styles = """
    <style>
        body { font-family: "Inter", sans-serif; background-color: #f3f3f3; margin: 0; padding: 0; }
        .email-container { max-width: 600px; margin: 20px auto; background-color: #fff; border-radius: 8px; 
                        overflow: hidden; box-shadow: 0 2px 6px rgba(0, 0, 0, 0.1); }
        .header { background-color: #f3f3f3; text-align: center; padding: 20px; }
        .header a { font-size: 32px; font-weight: bold; padding: 15px; margin: 5px 0; }
        .content { padding: 20px; color: #333; }
        .info { background-color: #f9f9f9; padding: 15px; border-left: 4px solid #028dff; margin: 20px 0; 
                border-radius: 4px; }
        .code { text-align: center; font-size: 32px; font-weight: bold; background-color: #f9f9f9; padding: 15px; 
                border: 1px solid #ddd; border-radius: 5px; margin: 20px 0; }
        .footer { background-color: #f3f3f3; padding: 10px; text-align: center; font-size: 12px; color: #777; }
        .footer a { color: #6441a5; text-decoration: none; }
    </style>
    """

    if email_type == "activation_kod":
        content = f"""
        <p>Здравствуйте!</p>
        <p>Кто-то пытается войти в ErespondentN используя вашу электронную почту.</p>
        <div class="info">
            {f'<p><strong>Расположение:</strong> {location}</p>' if location else ''}
            {f'<p><strong>Устройство:</strong> {device}</p>' if device else ''}
            {f'<p><strong>Браузер:</strong> {browser}</p>' if browser else ''}
            {f'<p><strong>IP-адрес:</strong> {ip_address}</p>' if ip_address else ''}
        </div>
        <p>Чтобы активировать вход, введите следующий код:</p>
        <div class="code">{message_body}</div>
        <p class="note">Обратите внимание, что срок действия этого кода истекает через 15 минут.</p>
        """
    elif email_type == "new_pass":
        content = f"""
        <p>Здравствуйте!</p>
        <p>Кто-то пытается изменить пароль в ErespondentN используя вашу электронную почту.</p>
        <div class="info">
            {f'<p><strong>Расположение:</strong> {location}</p>' if location else ''}
            {f'<p><strong>Устройство:</strong> {device}</p>' if device else ''}
            {f'<p><strong>Браузер:</strong> {browser}</p>' if browser else ''}
            {f'<p><strong>IP-адрес:</strong> {ip_address}</p>' if ip_address else ''}
        </div>
        <p>Вот ваш новый пароль:</p>
        <div class="code">{message_body}</div>
        <p>При желании его можно сменить в профиле пользователя, для этого требуется перейти:</p>
        <ul>
            <li>Личный кабинет →</li>
            <li>Параметры профиля →</li>
            <li>Безопасность</li>
        </ul>
        """
    elif email_type == "to_admin":
        content = f"""   
            <p style="margin: 0 0 18px 0; font-size: 16px; line-height: 1.6; color: #2c3e50; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">Здравствуйте!</p>
            <p style="margin: 0 0 25px 0; font-size: 16px; line-height: 1.6; color: #2c3e50; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">Сообщение администратору.</p>
            <p style="margin: 0 0 12px 0; font-size: 16px; line-height: 1.6; color: #2c3e50; font-weight: 600; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">Сообщение:</p>
            <div style="
                background: linear-gradient(120deg, rgba(67, 97, 238, 0.08) 0%, rgba(76, 201, 240, 0.08) 100%);
                border-left: 4px solid #4361ee;
                border-radius: 0 8px 8px 0;
                padding: 20px 25px;
                margin: 15px 0 25px 0;
                font-family: 'SF Mono', Monaco, 'Cascadia Code', 'Roboto Mono', monospace;
                font-size: 14.5px;
                line-height: 1.7;
                color: #2d3748;
                white-space: pre-wrap;
                word-wrap: break-word;
                overflow-wrap: break-word;
                box-shadow: 0 2px 8px rgba(67, 97, 238, 0.08);
            ">{message_body}</div>
        """
    elif email_type == "to_recipient":
        content = f"""   
            <p style="margin: 0 0 18px 0; font-size: 16px; line-height: 1.6; color: #2c3e50; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">Здравствуйте!</p>
            <p style="margin: 0 0 25px 0; font-size: 16px; line-height: 1.6; color: #2c3e50; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">Вам поступило сообщение от администратора.</p>
            <p style="margin: 0 0 12px 0; font-size: 16px; line-height: 1.6; color: #2c3e50; font-weight: 600; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">Сообщение:</p>
            <div style="
                background: linear-gradient(120deg, rgba(67, 97, 238, 0.08) 0%, rgba(76, 201, 240, 0.08) 100%);
                border-left: 4px solid #4361ee;
                border-radius: 0 8px 8px 0;
                padding: 20px 25px;
                margin: 15px 0 25px 0;
                font-family: 'SF Mono', Monaco, 'Cascadia Code', 'Roboto Mono', monospace;
                font-size: 14.5px;
                line-height: 1.7;
                color: #2d3748;
                white-space: pre-wrap;
                word-wrap: break-word;
                overflow-wrap: break-word;
                box-shadow: 0 2px 8px rgba(67, 97, 238, 0.08);
            ">{message_body}</div>
        """
    else:
        content = f"""
        <p>Здравствуйте!</p>
        <p>Сообщение об изменении статуса отчета.</p>
        <p>Статус отчета изменен на:</p>
        <div class="code">{message_body}</div>
        """

    html_template = f"""
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@100..900&display=swap" rel="stylesheet">
        {base_styles}
    </head>
    <body>
        <div class="email-container">
            
            <div class="content">
                {content}
            </div>
            <div class="footer">
                <p>Дополнительную информацию можно найти <a href="https://erespondentn.energoeffect.gov.by/FAQ">здесь</a>.</p>
                <p>Спасибо,<br>Служба поддержки ErespondentN</p>
            </div>
        </div>
    </body>
    </html>
    """
    # <div class="header"><a>ErespondentN</a></div>
    
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()

    try: 
        server.login(sender, password)
        
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = recipient_email
        msg["Subject"] = "Уведомление от ErespondentN"
        
        msg.attach(MIMEText(html_template, "html"))

        server.sendmail(sender, recipient_email, msg.as_string())
        
        # print("Письмо успешно отправлено")
        return "Email sent successfully"
    except Exception as _ex:
        # print(f"Ошибка при отправке письма: {_ex}")
        return f"{_ex}\nCheck log ..."
    finally:
        server.quit()

def send_activation_email(email):
    activation_kod = gener_password()
    session['activation_code'] = activation_kod
    send_email(activation_kod, email, 'activation_kod')

def gener_password():
    length=5
    characters = string.digits
    password = ''.join(random.choice(characters) for _ in range(length))
    return password

def parse_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return None

@auth.route('/login', methods=['POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False

        if email and password:
            user = User.query.filter(func.lower(User.email) == func.lower(email)).first()
            if user:
                if check_password_hash(user.password, password):
                    login_user(user, remember=remember)
                    session_token = create_user_session(user.id)
                    resp = make_response(redirect(url_for('views.account')))
                    resp.set_cookie('session_token', session_token, httponly=True)
                    flash('Авторизация прошла успешно.', 'success')
                    return resp
            flash('Неправильный email или пароль.', 'error')
        else:
            flash('Введите данные для авторизации.', 'error')

    return redirect(url_for('views.login'))

@auth.route('/logout')
@login_required
def logout():
    token = request.cookies.get('session_token')
    if token:
        session_obj = UserSession.query.filter_by(session_token=token).first()
        if session_obj:
            db.session.delete(session_obj)
            db.session.commit()

    response = make_response(redirect(url_for('views.login')))
    response.delete_cookie('session_token')
    logout_user()

    flash('Выполнен выход из аккаунта.', 'success')
    return response

@auth.route('/sign', methods=['POST'])
def sign():
    if request.method == 'POST':
        email = request.form.get('email')
        password1 = request.form.get('password1')
        password2 = request.form.get('password2')
        if email and password1:
            if User.query.filter(func.lower(User.email) == func.lower(email)).first():
                flash('Пользователь с таким email уже существует.', 'error')
            elif not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
                flash('Некорректный адрес электронной почты.', 'error') 
            elif password1 != password2:
                flash('Ошибка в подтверждении пароля.', 'error') 
            else:      
                session['temp_user'] = {
                    'email': email,
                    'password': generate_password_hash(password1)
                }
                session.permanent = True
                send_activation_email(email) 
                flash('Регистрация прошла успешно! Проверьте свою почту для активации аккаунта.', 'success')
                return redirect(url_for('views.kod'))
        else:
            flash('Введите данные для регистрации.', 'error')    
    return redirect(url_for('views.sign'))

@auth.route('/kod', methods=['POST'])
def kod():
    if request.method == 'POST':
        input_code = ''.join([
            request.form.get(f'activation_code_{i}', '') for i in range(5)
        ])
        if input_code == session.get('activation_code'):
            new_user = User(
                email=session['temp_user']['email'],
                password=session['temp_user']['password']
            )
            db.session.add(new_user)
            db.session.commit()
            session.pop('temp_user', None)
            session.pop('activation_code', None)
            flash('Аккаунт успешно активирован!', 'success')
            return redirect(url_for('views.login'))
        else:
            flash('Некорректный код активации.', 'error')
    return redirect(url_for('views.kod'))

@auth.route('/resend-code', methods=['POST'])
def resend_code():
    email = session.get('temp_user', {}).get('email')
    if email:
        new_activation_code = gener_password()
        session['activation_code'] = new_activation_code
        send_email(new_activation_code, email, 'activation_kod')
        return jsonify({'status': 'success', 'message': 'Код активации отправлен повторно!'})
    else:
        return jsonify({'status': 'error', 'message': 'Не удалось отправить код повторно.'}), 400

@auth.route('/sessions/clear-others', methods=['POST'])
@login_required
@session_required
def clear_other_sessions():
    current_token = get_session_token_from_cookie()
    other_sessions = UserSession.query.filter(
        UserSession.user_id == current_user.id,
        UserSession.session_token != current_token
    )

    count = other_sessions.count()

    if count == 0:
        flash('Нет других активных сессий.', 'error')
    else:
        other_sessions.delete(synchronize_session=False)
        db.session.commit()
        flash(f'Удалено {count} других сессий.', 'success')
    return redirect(url_for('views.profile_session'))

@auth.route('/add-personal-parametrs', methods=['POST'])
@login_required 
@session_required
def add_personal_parametrs():
    if request.method == 'POST':
        name = request.form.get('name_common', '').strip()
        second_name = request.form.get('second_name_common', '').strip()
        patronymic = request.form.get('patronymic_common', '').strip()
        telephone = request.form.get('telephone_common', '').strip()
        full_name = request.form.get('full_name_common', '').strip()

        if not all([name, second_name, telephone]):
            flash('Заполните все обязательные поля.', 'error')
            return redirect(url_for('views.profile_common'))

        fio = f"{second_name} {name} {patronymic}".strip()
        current_user.fio = fio
        db.session.commit()


        existing_telephone = User.query.filter(User.id != current_user.id, User.telephone == telephone).first()
        if existing_telephone:
            flash('Пользователь с таким номером телефона уже существует.', 'error')
            return redirect(url_for('views.profile_common'))
        current_user.telephone = telephone
        db.session.commit()
        
        if not all([full_name]):
            flash('Выберите предприятие.', 'error')
            return redirect(url_for('views.profile_common'))
        
        organiz_full_name = Organization.query.filter_by(full_name=full_name).first()
        if not organiz_full_name:
            flash('Организация не найдена.', 'error')
            return redirect(url_for('views.profile_common'))
        db.session.commit()

        existing_userOrg = User.query.filter_by(organization_id=organiz_full_name.id).first()
        if existing_userOrg and existing_userOrg.id != current_user.id:
            flash('Аккаунт с такой организацией уже существует.', 'error')
            return redirect(url_for('views.profile_common'))
        current_user.organization_id = organiz_full_name.id

        db.session.commit()
        flash('Данные успешно обновлены.', 'success')
        
    return redirect(url_for('views.profile_common'))

@auth.route('/profile/password', methods=['POST'])
@login_required
def profile_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        conf_new_password = request.form.get('conf_new_password')
        
        if not (old_password and new_password and conf_new_password):
            flash('Введите все поля для смены пароля.', 'error')
            return redirect(url_for('views.profile_password'))

        if not check_password_hash(current_user.password, old_password):
            flash('Неправильный старый пароль.', 'error')
            return redirect(url_for('views.profile_password'))

        if new_password != conf_new_password:
            flash('При подтверждении пароля произошла ошибка.', 'error')
            return redirect(url_for('views.profile_password'))

        user = User.query.filter_by(id=current_user.id).first()
        user.password = generate_password_hash(new_password)
        db.session.commit()

        # send_email(new_password, current_user.email, 'new_pass')

        UserSession.query.filter_by(user_id=current_user.id).delete()
        db.session.commit()

        logout_user()
        session.clear()

        response = make_response(redirect(url_for('auth.login')))
        response.delete_cookie('session_token')
        flash('Пароль изменён. Выполнен выход из всех устройств.', 'success')
        return response
    return redirect(url_for('views.profile_password'))

@auth.route('/relod-password', methods=['POST'])
def relod_password():
    email = request.form.get('email_relod')
    if not email:
        flash('Пожалуйста, введите свой email, затем нажмите «Забыли пароль?» для восстановления доступа.', 'error')
        return redirect(url_for('views.login'))
    user = User.query.filter(func.lower(User.email) == func.lower(email)).first()
    if user:
        new_password = gener_password()
        hashed_password = generate_password_hash(new_password)

        user_agent_string = request.headers.get('User-Agent')
        ip_address, location, os, browser = get_location_info(user_agent_string)
        
        send_email(new_password, email, 'new_pass', location=location, device=os, browser=browser, ip_address=ip_address)
        flash('Новый пароль был отправлен вам на email.', 'success')
        user.password = hashed_password
        db.session.commit()
        return redirect(url_for('views.login'))
    else:
        flash('Пользователя с таким email не существует.', 'error')
        return redirect(url_for('views.login'))

@auth.route('/create-new-report', methods=['POST'])
@login_required 
@session_required
def create_new_report():
    if request.method == 'POST':    
        organization_name = request.form.get('modal_organization_name')
        okpo = request.form.get('modal_organization_okpo')
        year =  parse_int(request.form.get('modal_report_year'))
        quarter =  parse_int(request.form.get('modal_report_quarter'))

        organization = Organization.query.filter_by(
            okpo = okpo, 
            full_name=organization_name).first()
        
        has_report = Report.query.filter_by(
            org_id=organization.id, 
            year = year, 
            quarter=quarter,
            user_id = current_user.id).first()
        
        if not has_report:
            new_report = Report(
                okpo=organization.okpo,
                org_id=organization.id,
                year=year,
                quarter=quarter,
                user_id = current_user.id
            )
            db.session.add(new_report)
            db.session.commit()
            new_version_report = Version_report(
                begin_time = current_utc_time(), 
                status = "Заполнение",
                fio = current_user.fio,
                telephone = current_user.telephone,
                email = current_user.email,
                report=new_report
            )
            db.session.add(new_version_report)
            db.session.commit() 

            sections = Sections.query.filter_by(id_version=new_version_report.id).all()
            if not sections:
                id = new_version_report.id
                sections_data = [
                    (id, 332, 9100, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 329, 9010, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 326, 9001, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
      
                    (id, 334, 9100, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 331, 9010, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 328, 9001, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                
                    (id, 333, 9100, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 330, 9010, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id, 327, 9001, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
        
                ]
                for data in sections_data:
                    section = Sections(
                        id_version=data[0],
                        id_product=data[1],
                        code_product=data[2],
                        section_number=data[3],
                        Oked=data[4],
                        produced=data[5],
                        Consumed_Quota=data[6],
                        Consumed_Fact=data[7],
                        Consumed_Total_Quota=data[8],
                        Consumed_Total_Fact=data[9],
                        total_differents=data[10],
                        note=data[11]
                    )
                    db.session.add(section)
                db.session.commit()
            flash('Добавлен новый отчет.', 'success')
        else:
            flash(f'Отчет {year} года {quarter} квартала уже существует.', 'error')
    return redirect(url_for('views.report_area'))

@auth.route('/update-report', methods=['POST'])
@login_required 
@session_required
def update_report():
    if request.method == 'POST':
        id = request.form.get('modal_report_id')
        okpo = request.form.get('modal_report_okpo')
        year = request.form.get('modal_change_report_year')
        quarter = request.form.get('modal_change_report_quarter')   
        current_report = Report.query.filter_by(id=id).first()
        versions = Version_report.query.filter_by(report_id=id).all()

        if current_report:
            existing_report = Report.query.filter_by(
                okpo=okpo,
                year=year,
                quarter=quarter
            ).first()

            sent_version_exists = any(version.status == 'Отправлен' for version in versions)
            if sent_version_exists:
                flash('После отправки изменение отчета недоступно.', 'error')
                return redirect(url_for('views.report_area'))
            
            confirmed_version_exists = any(version.status == 'Одобрен' for version in versions)
            if confirmed_version_exists:
                flash('Одобренные отчеты не подлежат редактированию.', 'error')
                return redirect(url_for('views.report_area'))
            
            if existing_report and existing_report.id != id:
                flash('Отчет с таким годом и кварталом уже существует.', 'error')
                return redirect(url_for('views.report_area'))

            current_report.okpo = okpo
            current_report.year = year
            current_report.quarter = quarter
            db.session.commit()
            flash('Параметры обновлены.', 'success')
        else:
            flash('Отчет не найден.', 'error')

        return redirect(url_for('views.report_area'))
    
@auth.route('/сopy-report', methods=['POST'])
@login_required 
@session_required
def сopy_report():
    if request.method == 'POST':
        coppy_report_id = request.form.get('coppy_report_id')
        new_organization_name = request.form.get('coppy_organization_name')
        new_organization_okpo = request.form.get('coppy_organization_okpo')
        new_report_year = request.form.get('coppy_report_year')
        new_report_quarter = request.form.get('coppy_report_quarter')

        current_report_version = Version_report.query.filter_by(
            report_id = coppy_report_id).first()
        
        current_sections = Sections.query.filter_by(
            id_version=current_report_version.id).all()

        current_org = Organization.query.filter_by(
            full_name = new_organization_name, 
            okpo = new_organization_okpo).first()

        proverka_report = Report.query.filter_by(
            org_id=current_org.id, 
            year = new_report_year, 
            quarter=new_report_quarter, 
            okpo = current_org.okpo).first()
        
        if not proverka_report:
            new_report = Report(
                okpo=current_org.okpo,
                org_id= current_org.id,
                year=new_report_year,
                quarter=new_report_quarter,
                user_id = current_user.id
            )
            db.session.add(new_report)
            db.session.commit()
            new_version = Version_report(
                status = "Заполнение",
                fio = current_user.fio,
                telephone = current_user.telephone,
                email = current_user.email,
                report=new_report
            )
            db.session.add(new_version)
            db.session.commit()
            for section in current_sections:
                new_section = Sections(
                    id_version=new_version.id,
                    id_product=section.id_product,
                    code_product=section.code_product,
                    section_number=section.section_number,
                    Oked=section.Oked,
                    produced=section.produced,
                    Consumed_Quota=section.Consumed_Quota,
                    Consumed_Fact=section.Consumed_Fact,
                    Consumed_Total_Quota=section.Consumed_Total_Quota,
                    Consumed_Total_Fact=section.Consumed_Total_Fact,
                    total_differents=section.total_differents,
                    note=section.note
                )
                db.session.add(new_section)
            db.session.commit()
            flash('Отчет успешно скопирован.', 'success')
        else:
            flash('Отчет с таким годом и квараталом уже существует.','error')
        return redirect(url_for('views.report_area'))

@auth.route('/delete-report/<report_id>', methods=['POST'])
@login_required 
@session_required
def delete_report(report_id):
    if request.method == 'POST':
        current_report = Report.query.filter_by(id = report_id).first()
        versions = Version_report.query.filter_by(report_id = report_id).all()
        tickets = Ticket.query.filter_by(version_report_id = report_id).all()  
        if current_report:   
            sent_version_exists = any(version.status == 'Отправлен' for version in versions)
            if sent_version_exists:
                flash('Отправленный отчет не подлежит удалению.', 'error')
                return redirect(url_for('views.report_area'))  
            confirmed_version_exists = any(version.status == 'Одобрен' for version in versions)
            if confirmed_version_exists:
                flash('Данный отчет не подлежит удалению.', 'error')
                return redirect(url_for('views.report_area'))
            for ticket in tickets:
                db.session.delete(ticket)        
            for version in versions:
                sections = Sections.query.filter_by(id_version = version.id).all()
                for section in sections:
                    db.session.delete(section)
                db.session.delete(version)
            db.session.delete(current_report)
            db.session.commit()
            flash('Отчет удален.', 'success')
        return redirect(url_for('views.report_area'))

@auth.route('/add-section-param', methods=['POST'])
@login_required 
@session_required
def add_section_param():
    if request.method == 'POST':
        current_version_id = request.form.get('current_version')
        add_id_product = request.form.get('add_id_product')
        oked = request.form.get('oked_add')
        produced = request.form.get('produced_add')
        Consumed_Quota = request.form.get('Consumed_Quota_add')
        Consumed_Fact = request.form.get('Consumed_Fact_add')
        Consumed_Total_Quota = request.form.get('Consumed_Total_Quota_add')
        Consumed_Total_Fact = request.form.get('Consumed_Total_Fact_add')
        note = request.form.get('note_add')
        section_number = request.form.get('section_number')
        
        produced = Decimal(produced) if produced else Decimal(0)
        Consumed_Quota = Decimal(Consumed_Quota) if Consumed_Quota else Decimal(0)
        Consumed_Fact = Decimal(Consumed_Fact) if Consumed_Fact else Decimal(0)
        Consumed_Total_Quota = Decimal(Consumed_Total_Quota) if Consumed_Total_Quota else Decimal(0)
        Consumed_Total_Fact = Decimal(Consumed_Total_Fact) if Consumed_Total_Fact else Decimal(0)

        current_product = DirProduct.query.filter_by(id=add_id_product).first()
        if not current_product:
            flash(f'Продукт не найден в справочнике.', 'error')
            return redirect(request.referrer)
        else:
            current_version = Version_report.query.filter_by(id=current_version_id).first()
            product_unit = DirUnit.query.filter_by(IdUnit=current_product.IdUnit).first() 
        
            if current_version:
                if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                    if current_product:
                        proverka_section = Sections.query.filter_by(
                            id_version=current_version_id, 
                            section_number=section_number, 
                            id_product=current_product.id
                        ).first()

                        if proverka_section and not note:
                            flash('Такой вид продукции уже существует — поле "Примечание" обязательно для заполнения.', 'error')
                            return redirect(request.referrer)

                        new_section = Sections(
                            id_version=current_version_id,
                            id_product=current_product.id,
                            code_product=current_product.CodeProduct,
                            section_number=section_number,
                            Oked=oked,
                            produced=produced,
                            Consumed_Quota=Consumed_Quota,
                            Consumed_Fact=Consumed_Fact,
                            Consumed_Total_Quota=Consumed_Total_Quota,
                            Consumed_Total_Fact=Consumed_Total_Fact,
                            total_differents=None,
                            note=note
                        )
                        db.session.add(new_section)
                        db.session.commit()
                        current_section = Sections.query.filter_by(
                            id=new_section.id, 
                            section_number=section_number
                        ).first()

                        if current_section.code_product == "7000":
                            current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                            db.session.commit()
                        else:
                            try:
                                if current_section.produced != 0:
                                    if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                        current_section.Consumed_Fact = round(
                                            (current_section.Consumed_Total_Fact / current_section.produced) * 100, 2)
                                    else:
                                        current_section.Consumed_Fact = round(
                                            (current_section.Consumed_Total_Fact / current_section.produced) * 1000, 2)
                                else:
                                    current_section.Consumed_Fact = 0

                                db.session.commit()
                                if current_section.Consumed_Quota != 0:
                                    if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                        current_section.Consumed_Total_Quota = round(
                                            (current_section.produced * current_section.Consumed_Quota) / 100, 2)
                                    else:
                                        current_section.Consumed_Total_Quota = round(
                                            (current_section.produced * current_section.Consumed_Quota) / 1000, 2)
                                else:
                                    current_section.Consumed_Total_Quota = 0
                                db.session.commit()
                                current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                                db.session.commit()
                            except InvalidOperation as e:
                                flash(f"Ошибка при вычислениях: {e}")

                        specific_codes = ['9001', '9010', '9100']
                        section9001 = Sections.query.filter_by(
                            id_version=current_version_id, 
                            section_number=section_number, 
                            code_product='9001'
                        ).first()
                        aggregated_values = db.session.query(
                            func.sum(Sections.Consumed_Total_Quota),
                            func.sum(Sections.Consumed_Total_Fact),
                            func.sum(Sections.total_differents)
                        ).filter(
                            Sections.id_version == current_version_id,
                            Sections.section_number == section_number,
                            ~Sections.code_product.in_(specific_codes)
                        ).first()

                        if section9001 and aggregated_values:
                            section9001.Consumed_Total_Quota, section9001.Consumed_Total_Fact, section9001.total_differents = aggregated_values
                            db.session.commit()

                        section9010 = Sections.query.filter_by(
                            id_version=current_version_id, section_number=section_number, code_product='9010').first()
                        section9100 = Sections.query.filter_by(
                            id_version=current_version_id, section_number=section_number, code_product='9100').first()

                        if section9100 and section9001 and section9010:
                            section9100.Consumed_Total_Quota = section9001.Consumed_Total_Quota + section9010.Consumed_Total_Quota
                            section9100.Consumed_Total_Fact = section9001.Consumed_Total_Fact + section9010.Consumed_Total_Fact
                            section9100.total_differents = section9001.total_differents + section9010.total_differents
                            db.session.commit()

                        if current_version:
                            current_version.change_time = current_utc_time()
                            current_version.status = "Заполнение"
                            current_version.sent_time = None
                            db.session.commit()

                        flash('Продукция была добавлена.', 'success')
                    else:
                        flash('Ошибка при обновлении.', 'error')   
                else:
                    flash('Редактирование отправленного/одобренного отчета недоступно.', 'error')
            else:
                flash('Версия не найдена.', 'error')
    return redirect(request.referrer)
        
@auth.route('/change-section', methods=['POST'])
@login_required 
@session_required
def change_section():
    if request.method == 'POST':
        id_version = request.form.get('current_version')
        id_section = request.form.get('id')
        produced = request.form.get('produced_change')
        Consumed_Quota = request.form.get('Consumed_Quota_change')     
        Consumed_Fact = request.form.get('Consumed_Fact_change')
        Consumed_Total_Quota = request.form.get('Consumed_Total_Quota_change')
        Consumed_Total_Fact = request.form.get('Consumed_Total_Fact_change')
        note = request.form.get('note_change')

        produced = Decimal(produced) if produced else Decimal(0)
        Consumed_Quota = Decimal(Consumed_Quota) if Consumed_Quota else Decimal(0)
        Consumed_Fact = Decimal(Consumed_Fact) if Consumed_Fact else Decimal(0)
        Consumed_Total_Quota = Decimal(Consumed_Total_Quota) if Consumed_Total_Quota else Decimal(0)
        Consumed_Total_Fact = Decimal(Consumed_Total_Fact) if Consumed_Total_Fact else Decimal(0)

        current_version = Version_report.query.filter_by(id=id_version).first() 
        current_section = Sections.query.filter_by(id=id_section).first()
        
        current_product = DirProduct.query.filter_by(id=current_section.id_product).first()
        product_unit = DirUnit.query.filter_by(IdUnit=current_product.IdUnit).first() if current_product else None
        

        if current_version:
            if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                if current_section:
                    if current_section.code_product == "7000":
                        current_section.Consumed_Total_Quota = Consumed_Total_Quota
                        current_section.Consumed_Total_Fact = Consumed_Total_Fact
                        current_section.note = note
                        db.session.commit()
                        
                        current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                        db.session.commit()
                    else:
                        current_section.produced = produced
                        current_section.Consumed_Quota = Consumed_Quota     
                        current_section.Consumed_Total_Fact = Consumed_Total_Fact
                        current_section.note = note 
                        db.session.commit()

                        try:
                            if current_section.produced != 0:
                                if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                    current_section.Consumed_Fact = round(
                                        (current_section.Consumed_Total_Fact / current_section.produced) * 100, 2)
                                else:
                                    current_section.Consumed_Fact = round(
                                        (current_section.Consumed_Total_Fact / current_section.produced) * 1000, 2)
                            else:
                                current_section.Consumed_Fact = 0

                            db.session.commit()
                            if current_section.Consumed_Quota != 0:
                                if product_unit and (product_unit.NameUnit == '%' or product_unit.NameUnit == '% (включая покупную)'):
                                    current_section.Consumed_Total_Quota = round(
                                        (current_section.produced * current_section.Consumed_Quota) / 100, 2)
                                else:
                                    current_section.Consumed_Total_Quota = round(
                                        (current_section.produced * current_section.Consumed_Quota) / 1000, 2)
                            else:
                                current_section.Consumed_Total_Quota = 0
                            db.session.commit()
                            current_section.total_differents = current_section.Consumed_Total_Fact - current_section.Consumed_Total_Quota
                            db.session.commit()
                        except InvalidOperation as e:
                            flash(f"Ошибка при вычислениях: {e}")



                    if current_version:
                        current_version.change_time = current_utc_time()
                        db.session.commit()

                    specific_codes = ['9001', '9010', '9100']
                    section9001 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product='9001').first()
                    aggregated_values = db.session.query(
                        func.sum(Sections.Consumed_Total_Quota),
                        func.sum(Sections.Consumed_Total_Fact),
                        func.sum(Sections.total_differents)
                    ).filter(
                        Sections.id_version == id_version,
                        Sections.section_number == current_section.section_number,
                        ~Sections.code_product.in_(specific_codes)
                    ).first()

                    if aggregated_values:
                        section9001.Consumed_Total_Quota = aggregated_values[0] or 0
                        section9001.Consumed_Total_Fact = aggregated_values[1] or 0
                        section9001.total_differents = aggregated_values[2] or 0
                        db.session.commit()

                    section9010 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product='9010').first()
                    section9100 = Sections.query.filter_by(id_version=id_version, section_number=current_section.section_number, code_product='9100').first()
                    if section9100 and section9001 and section9010:
                        section9100.Consumed_Total_Quota = (section9001.Consumed_Total_Quota or 0) + (section9010.Consumed_Total_Quota or 0)
                        section9100.Consumed_Total_Fact = (section9001.Consumed_Total_Fact or 0) + (section9010.Consumed_Total_Fact or 0)
                        section9100.total_differents = (section9001.total_differents or 0) + (section9010.total_differents or 0)
                        db.session.commit()
                    current_version.status = "Заполнение"
                    current_version.sent_time = None
                    db.session.commit()
                    flash('Параметры обновлены.', 'success')
                else:
                    flash('Ошибка при обновлении.', 'error')   
            else:
                flash('Редактирование отправленного/одобренного отчета недоступно.', 'error')
        else:
            flash('Версия не найдена.', 'error')
        if(current_section.section_number == 1):
            return redirect(url_for('views.report_section', report_type='fuel', id=id_version))
        elif(current_section.section_number == 2):
            return redirect(url_for('views.report_section', report_type='heat', id=id_version))
        elif(current_section.section_number == 3):
            return redirect(url_for('views.report_section', report_type='electro', id=id_version))

@auth.route('/remove_section/<id>', methods=['POST'])
@login_required 
@session_required
def remove_section(id):
    if request.method == 'POST':
        delete_section = Sections.query.filter_by(id=id).first()
        id_version = delete_section.id_version
        section_numberDELsection = delete_section.section_number
        current_version = Version_report.query.filter_by(id=id_version).first() 
        if current_version:
            if current_version.status != 'Отправлен' and current_version.status != 'Одобрен':
                if delete_section:
                    section9001 = Sections.query.filter_by(id_version=id_version, section_number = section_numberDELsection, code_product = '9001').first()
                    section9001.Consumed_Total_Quota -= delete_section.Consumed_Total_Quota
                    section9001.Consumed_Total_Fact -= delete_section.Consumed_Total_Fact
                    section9001.total_differents -= delete_section.total_differents

                    section9100 = Sections.query.filter_by(id_version=id_version, section_number = section_numberDELsection, code_product = '9100').first()
                    section9100.Consumed_Total_Quota -= delete_section.Consumed_Total_Quota
                    section9100.Consumed_Total_Fact -= delete_section.Consumed_Total_Fact
                    section9100.total_differents -= delete_section.total_differents

                    db.session.delete(delete_section)
                    db.session.commit()
                    flash('Продукция была удалена.', 'success')
                    if current_version:
                        current_version.change_time = current_utc_time()
                        db.session.commit()
                    current_version.status = "Заполнение"
                    current_version.sent_time = None
                    db.session.commit()
                else: 
                    flash('Ошибка при удалении', 'error')
            else:
                flash('Редактирование отправленного/одобренного отчета недоступно.', 'error')  
        else:
            flash('Версия не найдена.', 'error')

        match section_numberDELsection:
            case 1:
                return redirect(url_for('views.report_section', report_type='fuel', id=id_version))
            case 2:
                return redirect(url_for('views.report_section', report_type='heat', id=id_version))
            case 3:
                return redirect(url_for('views.report_section', report_type='electro', id=id_version))
            case _:
                return 404

def control_func(id):
    current_version = Version_report.query.filter_by(id=id).first()
    if not current_version:
        flash('Версия отчета не найдена.', 'error')
        return redirect(url_for('views.report_area'))

    id_version = current_version.id


    sections = {
        'fuel': Sections.query.filter_by(id_version=id, section_number=1, code_product='9010').first(),
        'heat': Sections.query.filter_by(id_version=id, section_number=2, code_product='9010').first(),
        'electro': Sections.query.filter_by(id_version=id, section_number=3, code_product='9010').first(),
    }

    if current_version.status == 'Заполнение':
        for key, section in sections.items():
            if section is None or not section.note:
                flash('Примечание с кодом строки "9010" обязательно для заполнения.', 'error')
                return redirect(url_for('views.report_section', report_type=key, id=id_version))

        current_version.status = 'Контроль пройден'
        db.session.commit()
        flash('Контроль пройден.', 'successful')
    else:
        flash('Контроль уже был пройден.', 'error')

    session['open_report_id'] = current_version.report_id
    return redirect(url_for('views.report_area'))

@auth.route('/control-version/<id>', methods=['POST'])
@login_required 
@session_required
def control_version(id):
    if request.method == 'POST':
        return control_func(id)

@auth.route('/agreed-version/<id>', methods=['POST'])
@login_required 
@session_required
def agreed_version(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        if current_version.status == 'Контроль пройден':     
            current_version.status = 'Согласовано'
            db.session.commit()
            flash('Отчет согласован.', 'successful')
        elif current_version.status == 'Согласовано': 
            flash('Отчет уже согласован.', 'succeful')
        else:
            flash('Необходимо пройти контроль.', 'error')
        session['open_report_id'] = current_version.report_id
        return redirect(url_for('views.report_area'))

@auth.route('/sent-version/<id>', methods=['POST'])
@login_required 
@session_required
def sent_version(id):
    if request.method == 'POST':
        uploaded_file = request.files.get('certificate')
        current_version = Version_report.query.filter_by(id=id).first()
        session['open_report_id'] = current_version.report_id
        
        if current_version.status == 'Отправлен':
            flash('Отчет уже отправлен.', 'error')
            return redirect(request.referrer)

        if current_version.status != 'Согласовано':
            flash('Необходимо согласовать.', 'error')
            return redirect(request.referrer)

        ALLOWED_EXTENSIONS = {'cer'}
        def allowed_file(filename):
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

        if not uploaded_file:
            flash('Файл сертификата обязателен.', 'error')
            return redirect(request.referrer)

        if not allowed_file(uploaded_file.filename):
            flash('Неверный формат файла. Загрузите файл в формате .cer.', 'error')
            return redirect(request.referrer)

        if not check_certificate_expiry(uploaded_file):
            flash('Срок действия сертификата истёк или файл некорректен.', 'error')
            return redirect(request.referrer)

        current_version.status = 'Отправлен'
        if current_version.sent_time is None:
            current_version.sent_time = current_utc_time()
        db.session.commit()

        flash('Отчет отправлен.', 'successful')
        return redirect(url_for('views.report_area'))

@auth.route('/cancel-sent-version/<id>', methods=['POST'])
@login_required 
@session_required
def cancle_sent_version(id):
    from .report.report import cancel_sending
    return cancel_sending(id)

@auth.route('/change-category-report', methods=['POST'])
@login_required 
@session_required
def change_category_report():
    action = request.form.get('action')
    report_id = request.form.get('reportId')
    status_itog = None
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=report_id).first()
        recipient_user = User.query.filter_by(email=current_version.email).first()

        if current_version is not None: 
            user = User.query.filter_by(email=current_version.email).first()
            if not current_version.hasNot and action != 'to_download':
                flash('Необходимо уточнить о каких ошибках идет речь.', 'error')
                return redirect(url_for('views.audit_report', id=current_version.id, modal='addCommentModal'))
            if action == 'not_viewed':
                status_itog = 'Отправлен'
            elif action == 'remarks':
                status_itog = 'Есть замечания'
            elif action == 'to_download':
                status_itog = 'Одобрен'
                ticket_message = Ticket(
                    note="Ошибок нет, отчет передан в следующую стадию проверки.",
                    luck=True,
                    version_report_id=current_version.id
                )
                db.session.add(ticket_message)
            elif action == 'to_delete':
                status_itog = 'Готов к удалению'
            else:
                flash('Неизвестное действие.', 'error')
                return redirect(request.referrer) 
            
            current_version.hasNot = False
            current_version.status = status_itog
            current_version.audit_time = current_utc_time()
            db.session.commit()

            user_message = Message(
                text = f"Статус вашего отчета №{current_version.id} был изменен.",
                sender_id = current_user.id,          
                recipient_id = recipient_user.id
            )
            db.session.add(user_message)
            db.session.commit()
    
            send_email(status_itog, user.email, 'change_status')

            flash(f'Статус вашего отчета №{current_version.id} был изменен.', 'success')
            return redirect(request.referrer) 
        else:
            flash('Отчет не найден.', 'error')
            return "Version not found", 404
        
@auth.route('/rollbackreport/<id>', methods=['POST'])
@login_required 
@session_required
def rollbackreport(id):
    if request.method == 'POST':
        current_version = Version_report.query.filter_by(id=id).first()
        recipient_user = User.query.filter_by(email=current_version.email).first()    
        if current_version:
            if current_version.status != 'Отправлен':
                if isinstance(current_version.audit_time, datetime):
                    audit_time = current_version.audit_time
                else:
                    audit_time = datetime.combine(current_version.audit_time, datetime.min.time())

                if audit_time + timedelta(days=30) <= current_utc_time():
                    flash('Прошло больше 30-го дня, статус отчета изменить нельзя.', 'error')
                else: 
                    current_version.status = "Отправлен"
                    current_version.hasNot = False
                 
                    user_message = Message(
                        text = f"Статус отчета был изменен.",
                        sender_id = recipient_user.id,    
                        recipient_id = recipient_user.id      
                    )
                    db.session.add(user_message)
                    db.session.commit()
                    
                    ticket_message = Ticket(
                        note="Возвращён в исходное состояние.",
                        luck=False,
                        version_report_id=current_version.id
                    )
                    db.session.add(ticket_message)
                    db.session.commit()
                    flash('Статус отчёта был изменён на «Непросмотренный».', 'success')
            else:
                flash('Статус отчёта уже установлен как «Непросмотренный».', 'error')
            return redirect(request.referrer)
        else:
            flash('Отчет не найден.', 'error')
    return redirect(request.referrer)

@auth.route('/send-comment', methods=['POST'])
@login_required 
@session_required
def send_comment():
    if request.method == 'POST':
        version_id = request.form.get('version_id')
        text = request.form.get('text')

        if not text or text.strip() == '':
            flash("Необходимо ввести текст.", "error")
            return redirect(request.referrer)
        
        cleaned_text = ' '.join(text.split())
        current_version = Version_report.query.filter_by(id=version_id).first()

        if current_version:
            new_comment = Ticket(
                note = cleaned_text,
                version_report_id = current_version.id
            )
            db.session.add(new_comment)
            current_version.hasNot = True
            db.session.commit()
            flash('Комментарий создан.', 'success')
        else:
            flash('Отчет не найден.', 'error')
        return redirect(url_for('views.audit_report', id = version_id))
    
def generate_excel_report(version_id):
    current_report = Report.query.filter_by(id=version_id).first()
    sections1 = Sections.query.filter_by(id_version=version_id, section_number=1).order_by(desc(Sections.id)).all()
    sections2 = Sections.query.filter_by(id_version=version_id, section_number=2).order_by(desc(Sections.id)).all()
    sections3 = Sections.query.filter_by(id_version=version_id, section_number=3).order_by(desc(Sections.id)).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    regular_font_9 = Font(name="Times New Roman", size=9)
    regular_font_9_italic = Font(name="Times New Roman", size=9, italic=True)
    regular_font_10 = Font(name="Times New Roman", size=10)
    regular_font_10_italic = Font(name="Times New Roman", size=10, italic=True)
    
    regular_font_11 = Font(name="Times New Roman", size=11)
    regular_font_11_italic = Font(name="Times New Roman", size=11, italic=True)
    
    regular_font_13 = Font(name="Times New Roman", size=13)
    bold_font_10 = Font(name="Times New Roman", size=10, bold=True)
    bold_font_11 = Font(name="Times New Roman", size=11, bold=True)
    bold_font_13 = Font(name="Times New Roman", size=13, bold=True)
    
    vertical_text = Alignment(horizontal="center", vertical="bottom", textRotation=90, wrap_text=True)
    top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    bottom = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    bottom_left = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_bottom = Side(border_style="thin", color="000000")
    bottom_border = Border(bottom=thin_bottom)


    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def set_cell(ws, row_start, col_start, row_end=None, col_end=None, text="", 
                        font=regular_font_10, row_height=None, alignment=left,
                        merge_direction='both'):  # 'horizontal', 'vertical', 'both', 'none'
        """
        merge_direction: 
        - 'horizontal': объединить только по столбцам (row_end игнорируется)
        - 'vertical': объединить только по строкам (col_end игнорируется)
        - 'both': объединить и по строкам и по столбцам
        - 'none': не объединять
        """
        if merge_direction == 'horizontal':
            row_end = row_start
            if col_end is None:
                col_end = col_start
        elif merge_direction == 'vertical':
            col_end = col_start
            if row_end is None:
                row_end = row_start
        elif merge_direction == 'both':
            if row_end is None:
                row_end = row_start
            if col_end is None:
                col_end = col_start
        else:
            row_end = row_start
            col_end = col_start
        
        if row_start != row_end or col_start != col_end:
            ws.merge_cells(start_row=row_start, start_column=col_start, 
                        end_row=row_end, end_column=col_end)
        
        cell = ws.cell(row=row_start, column=col_start)
        cell.value = text
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        
        if row_height is not None:
            for row in range(row_start, row_end + 1):
                ws.row_dimensions[row].height = row_height
        
        return cell
    
    def page_setttings(ws, print_area):
        # ws.print_area = print_area      
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

    def get_org_name_by_okpo(okpo):
        dop_org_data = [
            ('Брестское областное управление', '100000001000'),
            ('Витебское областное управление', '200000002000'),
            ('Гомельское областное управление', '300000003000'),
            ('Гродненское областное управление', '400000004000'),
            ('Управление г. Минск', '500000005000'),
            ('Минское областное управление', '600000006000'),
            ('Могилевское областное управление', '700000007000'),
            ('Департамент по энергоэффективности', '800000008000'),
        ]
    
        if not okpo or len(str(okpo)) < 4:
            return ""  
        
        okpo_str = str(okpo)
        if len(okpo_str) >= 4:
            fourth_from_end = okpo_str[-4]  
        else:
            return ""
        
        for name, code in dop_org_data:
            if code.endswith(str(fourth_from_end) + "000"):
                return name
        
            return ""

    def title_list(wb, report):
        ws = wb.create_sheet("Титульный лист", 0)
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 15:
                ws.row_dimensions[row].height = 14
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15

        # ===============================
        # Блоки текста
        # ===============================
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center 
                        
            okpo = report.okpo
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"{org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"областное (городское) управление по надзору за рациональным использованием ТЭР"
                        
            ws.merge_cells("B5:F5")
            ws["B5"].value = org_text
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
            
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(министерство, концерн, государственный комитет)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        # title_first_sign()
        # title_second_sign()
        
        ws.merge_cells("A14:O15")
        ws["A14"].value = f"ВЕДОМСТВЕННАЯ ОТЧЕТНОСТЬ О НОРМАХ РАСХОДА И (ИЛИ) ПРЕДЕЛЬНЫХ УРОВНЯХ ПОТРЕБЛЕНИЯ ТОПЛИВНО-ЭНЕРГЕТИЧЕСКИХ РЕСУРСОВ ЗА {report.quarter} КВАРТАЛ {report.year} Г."
        ws["A14"].font = bold_font_13
        ws["A14"].alignment = center
                                     
        ws.merge_cells("B16:N17")
        ws["B16"].value = f"{report.organization.full_name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):  # B=2, N=15
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:L18")
        ws["D18"].value = "(наименование юридического лица)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                        
        ws.merge_cells("B20:N20")
        ws["B20"].value = f"ОКПО: {report.okpo}".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Респондент:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:G25")
        ws["E25"].value = "Email"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:G26")
        ws["E26"].value = "ФИО"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "Телефон"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    

        ws.merge_cells("J25:M25")  
        ws["J25"].value = f"{report.user.email}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = left
        
        ws.merge_cells("J26:M26")    
        ws["J26"].value = f"{report.user.fio}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = left    
        
        ws.merge_cells("J27:M27")  
        ws["J27"].value = f"{report.user.telephone}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = left      
                     
        page_setttings(ws, print_area = "A1:O27")
        
        return ws
    
    def fuel_half_xlsx(wb, report):
        ws = wb.create_sheet("Часть 1")
        # ===============================
        # Колонки и строки
        # ===============================
        columns = [("A", 30), 
                ("B", 10), 
                ("C", 10), 
                ("D", 10),
                ("E", 10), 
                ("F", 10), 
                ("G", 10), 
                ("H", 10), 
                ("I", 10)
                ]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21.75
            elif row == 4:
                ws.row_dimensions[row].height = 23.25
            elif row == 5:
                ws.row_dimensions[row].height = 19.5
            elif row == 6:
                ws.row_dimensions[row].height = 159
            else:
                ws.row_dimensions[row].height = 15

        ws.merge_cells("A1:R1")
        ws["A1"].value = "Часть 1. Топливо"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center

        ws.merge_cells("A2:R2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        ws.merge_cells("A3:A6"); ws["A3"].value = "№ п/п"
        ws.merge_cells("B3:B6"); ws["B3"].value = "Код основных направлений энергосбережения"
        ws.merge_cells("C3:C6"); ws["C3"].value = "Наименование мероприятия"
        ws.merge_cells("D3:D6"); ws["D3"].value = "Единицы измерения"
        ws.merge_cells("E3:E6"); ws["E3"].value = "Объем внедрения"
        ws.merge_cells("F3:G5"); ws["F3"].value = "Условно-годовой экономический эффект"
        ws["F6"].value = "т у.т."
        ws["G6"].value = "тыс. руб."
        ws.merge_cells("H3:H6"); ws["H3"].value = "Ожидаемый срок внедрения мероприятия, квартал"
        ws.merge_cells("I3:I6"); ws["I3"].value = "Ожидаемый экономический эффект от внедрения мероприятия в текущем году, т у.т."
        ws.merge_cells("J3:J6"); ws["J3"].value = "Срок окупаемости, лет"
        ws.merge_cells("K3:K6"); ws["K3"].value = "Объем финансирования, тыс. руб."
        ws.merge_cells("L3:R3"); ws["L3"].value = "источники финансирования, руб."
        ws.merge_cells("L4:O4"); ws["L4"].value = "бюджетные"
        ws.merge_cells("L5:L6"); ws["L5"].value = "республиканский бюджет на финансирование госпрограммы"
        ws.merge_cells("M5:M6"); ws["M5"].value = "республиканский бюджет"
        ws.merge_cells("N5:N6"); ws["N5"].value = "местный бюджет"
        ws.merge_cells("O5:O6"); ws["O5"].value = "другие"
        ws.merge_cells("P4:P6"); ws["P4"].value = "собственные средства организации"
        ws.merge_cells("Q4:Q6"); ws["Q4"].value = "кредиты банков, займы"
        ws.merge_cells("R4:R6"); ws["R4"].value = "иные"

        for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=18):
            for cell in row:
                if not cell.value:
                    continue
                if cell.coordinate in ("D3", "E3", "F6", "G6", "J3", "K3", "H3", "I3", "L5", "M5", "N5", "O5", "P4", "Q4", "R4"):
                    cell.alignment = vertical_text
                else:
                    cell.alignment = top
                cell.font = regular_font_10

        row_index = 7
        for col in range(1, 19):
            cell = ws.cell(row=row_index, column=col, value=col)
            cell.alignment = center
            cell.font = regular_font_10

        # local_econ_execes = (EconExec.query
        #     .join(EconMeasure)
        #     .join(Plan)
        #     .filter(Plan.id == plan.id, EconExec.is_local == True)
        #     .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
        #     .all()
        # )

        # non_local_econ_execes = (EconExec.query
        #     .join(EconMeasure)
        #     .join(Plan)
        #     .filter(Plan.id == plan.id, EconExec.is_local == False)
        #     .options(joinedload(EconExec.econ_measures).joinedload(EconMeasure.plan))
        #     .all()
        # )

        # def add_section(title, execs, start_number=1):
        #     nonlocal row_index
        #     row_index += 1
        #     ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=18)
        #     ws.cell(row=row_index, column=1, value=title).font = bold_font_10
        #     ws.cell(row=row_index, column=1).alignment = center
        #     sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
        #     sums = {col: 0 for col in sum_cols}
        #     for idx, econ in enumerate(execs, start=start_number):
        #         row_index += 1
        #         row = [
        #             idx,
        #             econ.econ_measures.direction.code if econ.econ_measures and econ.econ_measures.direction else "",
        #             econ.name if hasattr(econ, "name") else "",
        #             econ.econ_measures.direction.unit.name if econ.econ_measures and econ.econ_measures.direction and econ.econ_measures.direction.unit else "",
        #             econ.Volume if hasattr(econ, "Volume") else 0,
        #             econ.EffTut if hasattr(econ, "EffTut") else 0,
        #             econ.EffRub if hasattr(econ, "EffRub") else 0,
        #             econ.ExpectedQuarter if hasattr(econ, "ExpectedQuarter") else 0,
        #             econ.EffCurrYear if hasattr(econ, "EffCurrYear") else 0,
        #             econ.Payback if hasattr(econ, "Payback") else 0,
        #             econ.VolumeFin if hasattr(econ, "VolumeFin") else 0,
        #             econ.BudgetState if hasattr(econ, "BudgetState") else 0,
        #             econ.BudgetRep if hasattr(econ, "BudgetRep") else 0,
        #             econ.BudgetLoc if hasattr(econ, "BudgetLoc") else 0,
        #             econ.BudgetOther if hasattr(econ, "BudgetOther") else 0,
        #             econ.MoneyOwn if hasattr(econ, "MoneyOwn") else 0,
        #             econ.MoneyLoan if hasattr(econ, "MoneyLoan") else 0,
        #             econ.MoneyOther if hasattr(econ, "MoneyOther") else 0
        #         ]
        #         for col in sum_cols:
        #             try:
        #                 sums[col] += float(row[col-1])
        #             except (TypeError, ValueError):
        #                 pass
        #         ws.append(row)
        #         for col_idx in range(1, 19):
        #             cell = ws.cell(row=row_index, column=col_idx)
        #             cell.alignment = left if col_idx == 3 else center
        #             cell.font = regular_font_10
        #             if col_idx in [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]:
        #                 cell.number_format = '0.000'
        #     row_index += 1
            
        #     ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        #     ws.cell(row=row_index, column=1, value="Итого по разделу:").alignment = left
        #     ws.cell(row=row_index, column=1).font = regular_font_10_italic
            
        #     for col in sum_cols:
        #         cell = ws.cell(row=row_index, column=col, value=sums[col])
        #         cell.alignment = center
        #         cell.font = regular_font_10_italic
        #         cell.number_format = '0.000'
        #     return start_number + len(execs)

        # next_number = add_section("Раздел 2. Мероприятия по экономии топливно-энергетических ресурсов", non_local_econ_execes, 1)
        # # add_section("Раздел 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов", local_econ_execes, next_number)

        # row_index += 1
        
        # ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        # ws.cell(row=row_index, column=1, value="Всего по части 2:").font = bold_font_10
        # ws.cell(row=row_index, column=1).alignment = left
        
        # sum_cols = [5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18]
        # total_sums = {col: 0 for col in sum_cols}
        # for row in ws.iter_rows(min_row=6, max_row=row_index, min_col=3, max_col=18):
        #     if row[0].value == "Итого по разделу:":
        #         for col in sum_cols:
        #             val = ws.cell(row=row[0].row, column=col).value
        #             try:
        #                 total_sums[col] += float(val)
        #             except (TypeError, ValueError):
        #                 pass
        # for col in sum_cols:
        #     cell = ws.cell(row=row_index, column=col, value=total_sums[col])
        #     cell.alignment = center
        #     cell.font = regular_font_10_italic
        #     cell.number_format = '0.000'

        # quarters = [
        #     ("      январь–март", "jan_mar"),
        #     ("      январь–июнь", "jan_jun"),
        #     ("      январь–сентябрь", "jan_sep"),
        #     ("      январь–декабрь", "jan_dec")
        # ]
        # local_totals = get_cumulative_econ_metrics(plan.id, False)
        
        # non_local_totals = get_cumulative_econ_metrics(plan.id, False)
        # total_metrics = {
        #     'jan_mar_eff': local_totals['jan_mar']['eff_curr_year'] + non_local_totals['jan_mar']['eff_curr_year'],
        #     'jan_mar_vol': local_totals['jan_mar']['volume_fin'] + non_local_totals['jan_mar']['volume_fin'],
        #     'jan_jun_eff': local_totals['jan_jun']['eff_curr_year'] + non_local_totals['jan_jun']['eff_curr_year'],
        #     'jan_jun_vol': local_totals['jan_jun']['volume_fin'] + non_local_totals['jan_jun']['volume_fin'],
        #     'jan_sep_eff': local_totals['jan_sep']['eff_curr_year'] + non_local_totals['jan_sep']['eff_curr_year'],
        #     'jan_sep_vol': local_totals['jan_sep']['volume_fin'] + non_local_totals['jan_sep']['volume_fin'],
        #     'jan_dec_eff': local_totals['jan_dec']['eff_curr_year'] + non_local_totals['jan_dec']['eff_curr_year'],
        #     'jan_dec_vol': local_totals['jan_dec']['volume_fin'] + non_local_totals['jan_dec']['volume_fin']
        # }
        
        # total_metrics = {
        #     'jan_mar_eff': local_totals['jan_mar']['eff_curr_year'],
        #     'jan_mar_vol': local_totals['jan_mar']['volume_fin'],
        #     'jan_jun_eff': local_totals['jan_jun']['eff_curr_year'],
        #     'jan_jun_vol': local_totals['jan_jun']['volume_fin'],
        #     'jan_sep_eff': local_totals['jan_sep']['eff_curr_year'],
        #     'jan_sep_vol': local_totals['jan_sep']['volume_fin'],
        #     'jan_dec_eff': local_totals['jan_dec']['eff_curr_year'],
        #     'jan_dec_vol': local_totals['jan_dec']['volume_fin']
        # }



        for row in ws.iter_rows(min_row=3, max_row=row_index, min_col=1, max_col=18):
            for cell in row:
                cell.border = thin_border

        
        return ws


    def add_sheet(ws, sections, title, unit_header_one_text, unit_header_all_text, nubmerT, nameT):
        ws.title = title

        merged_cells = {
            'A2:I3': f"РАЗДЕЛ {nubmerT}. {nameT}",
            'A4:A5': "Наименование вида продукции (работ услуг)",
            'B4:B5': "Код строки",
            'C4:C5': "Код по ОКЭД",
            'D4:D5': "Единица измерения",
            'E4:E5': "Произведено продукции (работ, услуг) за отчетный период",
            'F4:G4': f"Израсходовано на единицу продукции (работы, услуги) за отчетный период, {unit_header_one_text}",
            'H4:I4': f"Израсходовано на всю произведенную продукцию (работу, услугу) за отчетный период, {unit_header_all_text}",
            'F5': "по утвержденной норме (предельному уровню)",
            'G5': "фактически",
            'H5': "по утвержденной норме (предельному уровню)",
            'I5': "фактически",
            'A6': "A",
            'B6': "Б",
            'C6': "В",
            'D6': "Г",
            'E6': "1",
            'F6': "2",
            'G6': "3",
            'H6': "4",
            'I6': "5"
        }

        font = Font(name='Times New Roman', size=12)
        alignment_header = Alignment(wrap_text=True, vertical='center', horizontal='center')
        alignment_data_left = Alignment(wrap_text=True, vertical='center', horizontal='left')
        alignment_data_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for cell_range, text in merged_cells.items():
            top_left_cell = cell_range.split(':')[0]
            ws[top_left_cell] = text
            ws[top_left_cell].font = font
            ws[top_left_cell].alignment = alignment_header
            ws[top_left_cell].border = border

        for cell_range in merged_cells.keys():
            ws.merge_cells(cell_range)

        column_widths = {
            'A': 30,
            'B': 10,
            'C': 10,
            'D': 12,
            'E': 12,
            'F': 18,
            'G': 10,
            'H': 18,
            'I': 10,
        }
        row_heights = {
            4: 65,
            5: 65,
            6: 20
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        row_index = 7 
        for section in sections:
            note_text = f' - ({section.note})' if section.note else ''
            product_name = f"{section.product.NameProduct} {note_text}"
            
            ws[f'A{row_index}'] = product_name
            ws[f'B{row_index}'] = section.code_product
            ws[f'C{row_index}'] = section.Oked
            ws[f'D{row_index}'] = section.product.unit.NameUnit
            ws[f'E{row_index}'] = section.produced
            ws[f'F{row_index}'] = section.Consumed_Quota
            ws[f'G{row_index}'] = section.Consumed_Fact
            ws[f'H{row_index}'] = section.Consumed_Total_Quota
            ws[f'I{row_index}'] = section.Consumed_Total_Fact

            
            for col in column_widths.keys():
                cell = ws[f'{col}{row_index}']
                cell.border = border
                cell.font = font
                
                
                if col == 'A':
                    cell.alignment = alignment_data_left
                else:
                    cell.alignment = alignment_data_center

            row_index += 1
            
        page_setttings(ws, print_area="A1:I20")
    
    
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    wb.active = title_list(wb, current_report)
    
    add_sheet(wb.create_sheet(), sections1, "Раздел 1", 'кг у.т.', 'т у.т.', 1, "ТОПЛИВО")         
    add_sheet(wb.create_sheet(), sections2, "Раздел 2", 'Мкал', 'Гкал', 2, "ТЕПЛОВАЯ ЭНЕРГИЯ")             
    add_sheet(wb.create_sheet(), sections3, "Раздел 3", 'кВтч', 'тыс.кВтч', 3, "ЭЛЕКТРИЧЕСКАЯ ЭНЕРГИЯ")  
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(file_stream, as_attachment=True, download_name=f'{current_report.okpo}_{current_report.year}_{current_report.quarter}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@auth.route('/export-table', methods=['POST'])
@login_required 
@session_required
def export_table():
    version_id = int(request.form.get('version_id'))
    return generate_excel_report(version_id)

@auth.route('/export-version/<id>', methods=['POST'])
@login_required 
@session_required
def export_version(id):
    return generate_excel_report(id)

@auth.route('/print-ticket/<int:id>', methods=['POST'])
@login_required 
@session_required
def print_ticket(id):
    if request.method == 'POST':
        ticket = Ticket.query.get(id)
        if ticket is None:
            flash("Квитанция не найдена", "error")
            return redirect(request.referrer)

        version_report = ticket.version_report
        report = version_report.report

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)

        font_path_bold = os.path.join(os.path.dirname(__file__), 'static', 'fonts', 'Montserrat-Bold.ttf')
        font_path_regular = os.path.join(os.path.dirname(__file__), 'static', 'fonts', 'Montserrat-Regular.ttf')
        pdfmetrics.registerFont(TTFont('MontserratBold', font_path_bold))
        pdfmetrics.registerFont(TTFont('MontserratRegular', font_path_regular))

        c.setFont("MontserratBold", 12)
        y_position = 730 

        data = {
            "Отчет по форме": "Сведения о нормах",
            "Период": f"Год: {report.year}; Квартал: {report.quarter}",
            "ОКПО предприятия": report.okpo,
            "Результат обработки": "Отчет принят в обработку" if ticket.luck else "Отчет не принят в обработку",
            "Дата обработки": ticket.begin_time.strftime("%Y-%m-%d"),
            "Причина отказа": " " if ticket.luck else ticket.note,
        }

        name_x = 100
        value_x = 250
        y_position -= 20

        for key, value in data.items():
            c.setFont("MontserratBold", 12)
            c.drawString(name_x, y_position, f"{key}:")
            c.setFont("MontserratRegular", 12)
            c.drawString(value_x, y_position, f"{value}")
            y_position -= 20

        c.save()
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{report.okpo}_{report.year}_{report.quarter}.pdf",
            mimetype="application/pdf"
        )

@auth.route('/exportDBF', methods=['POST'])
@login_required 
@session_required
def exportDBF():
    if request.method == 'POST':
        try:
            versions = get_approved_versions(request.form)
            if not versions:
                flash('Нет одобренных отчетов для выбранных фильтров.', 'error')
                return redirect(request.referrer)

            zip_buffer = create_dbf_zip(versions)
            return send_zip_file(zip_buffer)
        except Exception as e:
            flash(f'Ошибка при экспорте DBF: {str(e)}', 'error')
            return redirect(request.referrer)

def get_approved_versions(form_data):
    """Получает одобренные версии отчетов по фильтрам"""
    year_filter = form_data.get('year_filter')
    quarter_filter = form_data.get('quarter_filter')

    filters = []
    if year_filter:
        filters.append(Report.year == year_filter)
    if quarter_filter:
        filters.append(Report.quarter == quarter_filter)

    current_user_type = current_user.type
    okpo_digit = str(current_user.organization.okpo)[-4]

    query = Version_report.query.options(
        joinedload(Version_report.report),
        joinedload(Version_report.sections).joinedload(Sections.product)
    ).join(Report).filter(
        Version_report.status == "Одобрен",
        *filters
    )

    if not (current_user_type == "Администратор" or (current_user_type == "Аудитор" and okpo_digit == "8")):
        query = query.filter(
            func.substr(func.cast(Report.okpo, String), func.length(Report.okpo) - 3, 1) == okpo_digit
        )

    return query.all()

def create_dbf_zip(versions):
    """Создает ZIP-архив с DBF-файлами"""
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED) as zip_file:
        for version in versions:
            dbf_data = prepare_dbf_data(version)
            dbf_content = create_dbf_file(dbf_data, version.report)
            zip_file.writestr(
                f'{version.report.okpo}_{version.report.year}_{version.report.quarter}.dbf',
                dbf_content
            )
    
    zip_buffer.seek(0)
    return zip_buffer

def prepare_dbf_data(version):
    """Подготавливает данные для DBF-файла"""
    report = version.report
    sections = version.sections
    data = []
    
    for section in sections:
        product = getattr(section, 'product', None)
        row = create_dbf_row(report, version, section, product)
        data.append(row)
    
    df = pd.DataFrame(data)
    df = sort_and_format_dataframe(df)
    return df

def create_dbf_row(report, version, section, product):
    """Создает строку данных для DBF"""
    unit_code = str(product.unit.CodeUnit) if (product and hasattr(product, 'unit') and product.unit) else ''
    code_product_str = str(section.code_product) if section.code_product else ''
    
    return {
        'YEAR_': str(report.year) if report.year is not None else '',
        'KVARTAL': str(report.quarter) if report.quarter is not None else '',
        'IDPREDPR': str(report.okpo) if report.okpo is not None else '',
        'DATERECEIV': version.sent_time.strftime('%d.%m.%Y') if version.sent_time else '',
        'EXCEED': '0,000',
        'SECTIONNUM': str(section.section_number) if section.section_number is not None else '',
        'CODEPROD': code_product_str,
        'OKED': str(section.Oked) if section.Oked is not None else '',
        'PRODUCED': format_number(section.produced), 
        'CONSUMEDQ': format_number(section.Consumed_Quota),
        'CONSUMEDF': format_number(section.Consumed_Fact),
        'CONSUMEDQT': format_number(section.Consumed_Total_Quota),
        'CONSUMEDFT': format_number(section.Consumed_Total_Fact),
        'COMMENT1': str(section.note) if section.note is not None else '',
        'COMMENT2': '',
        'NAMEPROD': encode_cp866(product.NameProduct[:254]) if product and hasattr(product, 'NameProduct') and product.NameProduct else '',
        'CODEUNIT': unit_code[:20] if unit_code else '',
        'NAMEORG': encode_cp866(report.organization.full_name[:254]) if report.organization.full_name is not None else '',
        '_SECTIONNUM': int(section.section_number) if section.section_number else 0,
        '_CODEPROD': int(section.code_product) if section.code_product and section.code_product.isdigit() else 0,
        '_SORT_PRIORITY': {'9001': 0, '9010': 1, '9100': 2}.get(code_product_str, 3),
    }

def sort_and_format_dataframe(df):
    """Сортирует и форматирует DataFrame перед экспортом"""
    df = df.sort_values(by=['_SECTIONNUM', '_SORT_PRIORITY', '_CODEPROD'])
    df = df.drop(columns=['_SECTIONNUM', '_CODEPROD', '_SORT_PRIORITY'])
    df['INDX'] = range(1, len(df) + 1)
    df['INDX'] = df['INDX'].astype(str)
    return df[['INDX'] + [col for col in df.columns if col != 'INDX']]

def create_dbf_file(df, report):
    """Создает DBF-файл из DataFrame"""
    with NamedTemporaryFile(delete=False, suffix='.dbf') as temp_file:
        temp_filename = temp_file.name
        
        table = dbf.Table(
            temp_filename,
            'INDX C(10); YEAR_ C(20); KVARTAL C(20); IDPREDPR C(20); DATERECEIV C(20); '
            'EXCEED C(20); SECTIONNUM C(20); CODEPROD C(20); OKED C(20); PRODUCED C(20); '
            'CONSUMEDQ C(20); CONSUMEDF C(20); CONSUMEDQT C(20); CONSUMEDFT C(20); '
            'COMMENT1 C(200); COMMENT2 C(200); NAMEPROD C(200); CODEUNIT C(20); NAMEORG C(254); ',
            codepage='cp866'
        )
        
        table.open(mode=dbf.READ_WRITE)
        try:
            for _, row in df.iterrows():
                row_dict = row.to_dict()
                for key, value in row_dict.items():
                    if pd.isna(value) or value is None:
                        row_dict[key] = ''
                table.append(row_dict)
        finally:
            table.close()
        
        with open(temp_filename, 'rb') as f:
            return f.read()

def format_number(value, decimal_places=None):
    """Форматирует число с запятой в качестве десятичного разделителя"""
    if value is None or value == '':
        return '0'
    
    try:
        str_value = str(value).strip()
        str_value = str_value.replace('.', ',').replace(',', '.', 1) if '.' in str_value and ',' in str_value else str_value
        str_value = str_value.replace(',', '.')
        
        num = float(str_value)
        if decimal_places is None:
            if '.' in str_value:
                decimal_places = len(str_value.split('.')[1])
            else:
                decimal_places = 0
        
        if decimal_places == 0:
            return f"{int(round(num))}"
        else:
            formatted = f"{num:.{decimal_places}f}"
            return formatted.replace('.', ',')
            
    except (ValueError, TypeError):
        return '0,000'

def encode_cp866(text):
    """Кодирует текст в CP866"""
    return text.encode('cp866', 'replace').decode('cp866')

def send_zip_file(zip_buffer):
    """Отправляет ZIP-файл клиенту"""
    return Response(
        zip_buffer,
        mimetype='application/zip',
        headers={"Content-Disposition": "attachment;filename=reports_DBF.zip"}
    )


from collections import defaultdict
from lxml import etree
from collections import defaultdict

def create_xml_for_version(version):
    report = version.report
    organization = report.organization

    root = etree.Element("REPOPT_ROOT")
    report_el = etree.SubElement(root, "REPORT")

    etree.SubElement(report_el, "DESCRIPTION_REPORT", attrib={
        "CODE": str(report.okpo or ""),
        "NO": organization.full_name or "",
        "CODE_ENT": organization.full_name or "",
        "OKPO": str(report.okpo or ""),
        "UNP": str(organization.ynp or ""),
        "YEAR": str(report.year),
        "QUARTER": str(report.quarter),
        "PHASE": "1",
        "ISYEAR": "false",
        "TYPE": "2",
        "ISREADONLY": "false",
        "FIO": version.fio or "",
        "PHONE": version.telephone or "",
        "EMAIL": version.email or "",
        "SV": "1.2.7.2"
    })

    row_data_el = etree.SubElement(report_el, "ROW_DATA")
    sections_el = etree.SubElement(row_data_el, "SECTIONS")

    sections_by_number = defaultdict(list)
    for section in version.sections:
        sections_by_number[section.section_number].append(section)

    special_codes_order = ["9001", "9010", "9100"]

    for section_number, group in sections_by_number.items():
        for code in special_codes_order:
            for section in group:
                product = section.product
                if (product.CodeProduct or "") == code:
                    unit = product.unit
                    etree.SubElement(sections_el, "row", attrib={
                        "SectionNumber": str(section.section_number),
                        "SortIndex": code,
                        "CodeProduct": code,
                        "NameProduct": product.NameProduct or "",
                        "OKED": section.Oked or "",
                        "CodeUnit": unit.CodeUnit if unit else "",
                        "Produced": str(section.produced or 0),
                        "ConsumedQ": str(section.Consumed_Quota or 0),
                        "ConsumedF": str(section.Consumed_Fact or 0),
                        "ConsumedQT": str(section.Consumed_Total_Quota or 0),
                        "ConsumedFT": str(section.Consumed_Total_Fact or 0),
                        "Comment": section.note or ""
                    })

        sort_index_counter = 1
        for section in group:
            product = section.product
            code_product = product.CodeProduct or ""
            if code_product not in special_codes_order:
                unit = product.unit
                etree.SubElement(sections_el, "row", attrib={
                    "SectionNumber": str(section.section_number),
                    "SortIndex": str(sort_index_counter),
                    "CodeProduct": code_product,
                    "NameProduct": product.NameProduct or "",
                    "OKED": section.Oked or "",
                    "CodeUnit": unit.CodeUnit if unit else "",
                    "Produced": str(section.produced or 0),
                    "ConsumedQ": str(section.Consumed_Quota or 0),
                    "ConsumedF": str(section.Consumed_Fact or 0),
                    "ConsumedQT": str(section.Consumed_Total_Quota or 0),
                    "ConsumedFT": str(section.Consumed_Total_Fact or 0),
                    "Comment": section.note or ""
                })
                sort_index_counter += 1

    return etree.tostring(root, pretty_print=True, encoding='utf-8')

@auth.route('/exportXML', methods=['POST'])
@login_required 
@session_required
def exportXML():
    year_filter = request.form.get('year_filter')
    quarter_filter = request.form.get('quarter_filter')

    filters = []
    if year_filter:
        filters.append(Report.year == year_filter)
    if quarter_filter:
        filters.append(Report.quarter == quarter_filter)

    current_user_type = current_user.type
    okpo_digit = str(current_user.organization.okpo)[-4]

    if current_user_type == "Администратор" or (current_user_type == "Аудитор" and okpo_digit == "8"):
        versions = Version_report.query.options(
            joinedload(Version_report.report),
            joinedload(Version_report.sections).joinedload(Sections.product)
        ).join(Report).filter(
            Version_report.status == "Одобрен",
            *filters
        ).all()
    else:
        versions = Version_report.query.options(
            joinedload(Version_report.report),
            joinedload(Version_report.sections).joinedload(Sections.product)
        ).join(Report).filter(
            Version_report.status == "Одобрен",
            func.substr(func.cast(Report.okpo, String), func.length(Report.okpo) - 3, 1) == okpo_digit,
            *filters
        ).all()
    
    if not versions:
        flash('Нет одобренных отчетов для выбранных фильтров.', 'error')
        return redirect(request.referrer) 

    mem_zip = io.BytesIO()

    with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for version in versions:
            report = version.report
            xml_data = create_xml_for_version(version)
            file_name = f"{report.okpo}_{report.year}_{report.quarter}.xml"
            zipf.writestr(file_name, xml_data)
    mem_zip.seek(0)

    return send_file(
        mem_zip,
        as_attachment=True,
        download_name="reports_XML.zip",
        mimetype="application/zip"
    )

@auth.route('/sent_for_admin', methods=['POST'])
@login_required 
@session_required
def sent_for_admin():
    if request.method == 'POST':
        text = request.form.get('text')
        if text:
            admins = User.query.filter_by(type = "Администратор").all()
            if admins:
                for i in admins:
                    new_message = Message(
                        sender_id = current_user.id,
                        text = text,
                        recipient_id = i.id
                    )
                    db.session.add(new_message)
                db.session.commit()
                    
                flash('Сообщение отправлено.', 'succes')
                admin_user = os.getenv('adminemail3')
                if admin_user:
                    try:
                        send_email(text, admin_user, 'to_admin')
                    except Exception as e:
                        auth.logger.error(f"Ошибка отправки email: {str(e)}")
            else:
                flash('Администраторов нет.', 'error')
        else:
            flash('Нельзя отравить пустое сообщение.', 'error')
    return redirect(url_for('views.beginPage'))






def get_organizations_with_reports_excel_xlsx(year: int, quarter: int, statuses: list) -> bytes:
    status_filter = 'all_reports' if not statuses else statuses[0] if len(statuses) == 1 else 'all_reports'
    
    from .views import get_reports_by_status
    reports = get_reports_by_status(status_filter, year, quarter)
    
    if not reports:
        return None

    organizations_data = set()
    for report in reports:
        valid_versions = [v for v in report.versions if not statuses or v.status in statuses]
        if valid_versions:
            latest_version = max(valid_versions, key=lambda x: x.sent_time or datetime.min)
            organizations_data.add((
                report.organization.okpo,
                report.organization.full_name,
                latest_version.sent_time
            ))

    if not organizations_data:
        return None

    records = sorted(organizations_data, key=lambda x: x[0] or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Организации"

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="C6EFCE")
    title_font = Font(bold=True, size=12)
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_row = 3
    start_col = 2

    title_text = f"Список предприятий, представивших отчеты по форме Сведения о нормах в электронном виде на {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws.merge_cells(start_row=start_row - 1, start_column=start_col, end_row=start_row - 1, end_column=start_col + 3)
    title_cell = ws.cell(row=start_row - 1, column=start_col, value=title_text)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row - 1].height = 45

    for col in range(start_col, start_col + 4):
        cell = ws.cell(row=start_row - 1, column=col)
        cell.border = thin_border

    headers = ['Код предприятия (ОКПО)', 'Наименование предприятия', 'Дата поступления', 'Примечание']
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    for row_offset, (okpo, full_name, sent_time) in enumerate(records, start=1):
        row_data = [okpo or '', full_name or '', sent_time or '', '']
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(
                row=start_row + row_offset,
                column=start_col + col_offset,
                value=value
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            if col_offset == 2 and sent_time:
                cell.number_format = 'YYYY-MM-DD'

    for i in range(len(headers)):
        col_letter = get_column_letter(start_col + i)
        max_length = len(headers[i])
        for j in range(1, len(records) + 1):
            val = ws.cell(row=start_row + j, column=start_col + i).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@auth.route('/load_org_stat', methods=['POST'])
@login_required 
@session_required
def load_org_stat():
    year_filter = request.form.get('modal_report_year')
    quarter_filter = request.form.get('modal_report_quarter')

    if not year_filter or not quarter_filter:
        flash('Не указан год или квартал.', 'error')
        return redirect(request.referrer)

    if current_user.type not in ["Администратор", "Аудитор"]:
        flash('У вас нет доступа к отчетам.', 'error')
        return redirect(request.referrer)

    allowed_statuses = ["Отправлен", "Одобрен", "Есть замечания", "Готов к удалению"]
    file_data = get_organizations_with_reports_excel_xlsx(
        int(year_filter), int(quarter_filter), allowed_statuses
    )

    if not file_data:
        flash('Нет организаций с отправленными отчётами за выбранный период.', 'error')
        return redirect(request.referrer)

    response = make_response(file_data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=orgs_{year_filter}_{quarter_filter}.xlsx'
    return response

def create_test_users(userpassword):
    hashed_pass = generate_password_hash(userpassword)
    users_data = [
        ('testuser1BrestReg@gmail.com', 'Григорьев Антон Олегович', 'Пусто', hashed_pass, 3),
        ('testuser2BrestReg@gmail.com', 'Иванов Алексей Дмитриевич', 'Пусто', hashed_pass, 4),
        ('testuser1VitebskReg@gmail.com', 'Петров Николай Сергеевич', 'Пусто', hashed_pass, 1170),
        ('testuser2VitebskReg@gmail.com', 'Кузнецов Андрей Валерьевич', 'Пусто', hashed_pass, 1171),
        ('testuser1GomelReg@gmail.com', 'Фёдоров Дмитрий Иванович', 'Пусто', hashed_pass, 2229),
        ('testuser2GomelReg@gmail.com', 'Морозов Сергей Александрович', 'Пусто', hashed_pass, 2230),
        ('testuser1GrodnoReg@gmail.com', 'Васильев Игорь Михайлович', 'Пусто', hashed_pass, 3431),
        ('testuser2GrodnoReg@gmail.com', 'Новиков Евгений Павлович', 'Пусто', hashed_pass, 3432),
        ('testuser1MinskReg@gmail.com', 'Мельников Артём Владимирович', 'Пусто', hashed_pass, 4674),
        ('testuser2MinskReg@gmail.com', 'Орлов Роман Викторович', 'Пусто', hashed_pass, 4675),
        ('testuser1MogilevReg@gmail.com', 'Смирнов Константин Петрович', 'Пусто', hashed_pass, 6607),
        ('testuser2MogilevReg@gmail.com', 'Козлов Денис Сергеевич', 'Пусто', hashed_pass, 6608),
        ('testuser1Minsk@gmail.com', 'Зайцев Павел Андреевич', 'Пусто', hashed_pass, 7493),
        ('testuser2Minsk@gmail.com', 'Соловьёв Владимир Николаевич', 'Пусто', hashed_pass, 6899),
    ]

    current_time = datetime.utcnow()
    current_year = current_time.year
    current_quarter = 1

    created_version_ids = []

    for user_data in users_data:
        user = User.query.filter_by(email=user_data[0]).first()
        if not user:
            user = User(
                email=user_data[0],
                fio=user_data[1],
                telephone=user_data[2],
                password=user_data[3],
                organization_id=user_data[4],
            )
            db.session.add(user)
            db.session.flush()

        organization = Organization.query.get(user.organization_id)
        if not organization:
            continue

        # 2 отчёта: текущий год и следующий квартал
        for quarter in (current_quarter, current_quarter + 1):
            new_report = Report(
                okpo=organization.okpo,
                org_id=organization.id,
                year=current_year,
                quarter=quarter,
                user_id=user.id
            )
            db.session.add(new_report)
            db.session.flush()

            new_version_report = Version_report(
                begin_time=current_time,
                status="Заполнение",
                fio=user.fio,
                telephone=user.telephone,
                email=user.email,
                report=new_report
            )
            db.session.add(new_version_report)
            db.session.flush()

            created_version_ids.append(new_version_report.id)

            sections = Sections.query.filter_by(id_version=new_version_report.id).all()
            if not sections:
                id_version = new_version_report.id
                sections_data = [
                    (id_version, 326, 9100, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('33.00'), Decimal('411.00'), Decimal('378.00'), ''),
                    (id_version, 329, 9010, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), 'какая-то информация'),
                    (id_version, 332, 9001, 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('33.00'), Decimal('411.00'), Decimal('378.00'), ''),

                    (id_version, 6, '0026', 1, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('33.00'), Decimal('411.00'), Decimal('378.00'), ''),

                    (id_version, 327, 9100, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id_version, 330, 9010, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), 'какая-то информация'),
                    (id_version, 333, 9001, 2, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id_version, 328, 9100, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                    (id_version, 331, 9010, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), 'какая-то информация'),
                    (id_version, 334, 9001, 3, '', Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), ''),
                ]
                for data in sections_data:
                    section = Sections(
                        id_version=data[0],
                        id_product=data[1],
                        code_product=data[2],
                        section_number=data[3],
                        Oked=data[4],
                        produced=data[5],
                        Consumed_Quota=data[6],
                        Consumed_Fact=data[7],
                        Consumed_Total_Quota=data[8],
                        Consumed_Total_Fact=data[9],
                        total_differents=data[10],
                        note=data[11]
                    )
                    db.session.add(section)

    db.session.commit()

    versions_to_update = Version_report.query.filter(Version_report.id.in_(created_version_ids)).all()
    for version in versions_to_update:
        version.status = 'Отправлен'
        version.sent_time = current_utc_time()
    db.session.commit()

    return "Пользователи и два отчёта на каждый год созданы"

def delete_test_users():
    emails = [
        'testuser1BrestReg@gmail.com',
        'testuser2BrestReg@gmail.com',
        'testuser1VitebskReg@gmail.com',
        'testuser2VitebskReg@gmail.com',
        'testuser1GomelReg@gmail.com',
        'testuser2GomelReg@gmail.com',
        'testuser1GrodnoReg@gmail.com',
        'testuser2GrodnoReg@gmail.com',
        'testuser1MinskReg@gmail.com',
        'testuser2MinskReg@gmail.com',
        'testuser1MogilevReg@gmail.com',
        'testuser2MogilevReg@gmail.com',
        'testuser1Minsk@gmail.com',
        'testuser2Minsk@gmail.com',
    ]
    
    users = User.query.filter(User.email.in_(emails)).all()
    for user in users:
        db.session.delete(user)
    db.session.commit()
    return "Пользователи удалены"

@auth.route('/create-test-users', methods=['POST'])
@login_required 
@session_required
def create_test_users_rout():
    return create_test_users(userpassword = '1111')

@auth.route('/delete-test-users', methods=['POST'])
@login_required 
@session_required
def delete_test_users_rout():
    return delete_test_users()